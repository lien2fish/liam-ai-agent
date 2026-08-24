"""解析「流動資產管理」xlsx，輸出三品牌正規化庫存資料。

用法：python3 crm_unified/parse_inventory.py [xlsx路徑]
不帶參數則用 SRC 預設路徑。鑫酒藏另讀 WINE_SRC 的「總表」分頁。
輸出 JSON 到 stdout 供 sync_inventory.py 使用。
"""

import json
import re
import sys

import openpyxl

SRC = (
    "/Users/lien/Desktop/鉅鑫管理顧問/20260728流動資產管理─鑫酒藏、鑫海產、鑫茶坊.xlsx"
)

TAEL_TO_CATTY = {"1斤": 1.0, "4兩": 0.25, "2兩": 0.125}

WINE_SRC = "/Users/lien/Desktop/鉅鑫管理顧問/鑫酒藏庫存 2026.07.02.xlsx"

WINE_HEADERS = [
    "分類",
    "國家/產地",
    "品牌/酒莊",
    "酒種",
    "品種",
    "名稱/年份/容量",
    "備註",
    "鑫酒藏",
    "匠鑫私廚",
    "綠塔",
    "進貨價",
    "進貨總值",
    "產品定價",
    "批發價",
    "零售價",
]

WINE_LOCATIONS = ["鑫酒藏", "匠鑫私廚", "綠塔"]


def parse_price(raw):
    """'1000/台斤' -> (1000.0, '台斤')；'850' -> (850.0, '')"""
    if raw is None:
        return None, ""
    s = str(raw).strip()
    if not s:
        return None, ""
    m = re.match(r"^([\d,.]+)\s*/?\s*(.*)$", s)
    if not m:
        return None, s
    try:
        return float(m.group(1).replace(",", "")), m.group(2).strip()
    except ValueError:
        return None, s


def clean(v):
    return str(v).strip() if v is not None and str(v).strip() else ""


def parse_seafood(ws):
    items = []
    category = ""
    unit_hint = ""
    for r in ws.iter_rows(min_row=3, values_only=True):
        name = clean(r[1])
        if not name:
            continue
        if clean(r[0]):
            category = clean(r[0])
        if clean(r[3]):
            unit_hint = clean(r[3])
        price, price_unit = parse_price(r[5])
        qty = r[4] if isinstance(r[4], (int, float)) else 0
        items.append(
            {
                "品名": name,
                "產品種類": category,
                "分裝單位": unit_hint,
                "庫存數量": qty,
                "進價": price,
                "計價單位": price_unit,
                "備註": clean(r[7]),
            }
        )
    return items


def qty(v):
    """數量欄混著文字型數字（'0'、'12'），一律轉 int，轉不動當 0。"""
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return 0


def parse_wine(ws):
    header = [clean(v) for v in next(ws.iter_rows(max_row=1, values_only=True))]
    if header != WINE_HEADERS:
        raise ValueError(f"鑫酒藏總表欄位與預期不符：{header}")

    items = []
    for row_no, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        winery, name = clean(r[2]), clean(r[5])
        if not winery and not name:
            continue
        counts = [qty(r[7]), qty(r[8]), qty(r[9])]
        price = r[10] if isinstance(r[10], (int, float)) else None
        items.append(
            {
                "品名": f"{winery} {name}".strip() or f"（第{row_no}列）",
                "來源列": row_no,
                "分類": re.sub(r"[A-Z]+$", "", clean(r[0])),
                "國家": re.sub(r"\d+$", "", clean(r[1])),
                "酒莊": winery,
                "酒種": clean(r[3]),
                "品種": clean(r[4]),
                "存放位置": "+".join(
                    loc for loc, c in zip(WINE_LOCATIONS, counts) if c
                ),
                "庫存數量": counts[0],
                "匠鑫私廚": counts[1],
                "綠塔": counts[2],
                "進價": price,
                "庫存成本": counts[0] * price if price else None,
                "進貨總值": r[11] if isinstance(r[11], (int, float)) else None,
                "產品定價": r[12] if isinstance(r[12], (int, float)) else None,
                "批發價": r[13] if isinstance(r[13], (int, float)) else None,
                "零售價": r[14] if isinstance(r[14], (int, float)) else None,
                "備註": clean(r[6]),
            }
        )

    # 同一支酒在總表被登錄兩次（上方暫存區＋下方國別區），保留後出現的正式登錄
    unique = {}
    for item in items:
        unique[(item["酒莊"], item["品名"], item["酒種"], item["品種"])] = item
    dropped = [i for i in items if i is not unique.get((i["酒莊"], i["品名"], i["酒種"], i["品種"]))]
    if dropped:
        print(
            "鑫酒藏總表重複登錄，已略過："
            + "、".join(f"第{i['來源列']}列 {i['品名']}" for i in dropped),
            file=sys.stderr,
        )
    items = list(unique.values())

    def regroup():
        groups = {}
        for item in items:
            groups.setdefault(item["品名"], []).append(item)
        return {k: v for k, v in groups.items() if len(v) > 1}

    for name, group in regroup().items():
        variants = [f"{i['酒種']} {i['品種']}".strip() for i in group]
        if len(set(variants)) == len(group):
            for item, variant in zip(group, variants):
                item["品名"] = f"{name}（{variant}）"
    for name, group in regroup().items():
        for item in group:
            item["品名"] = f"{name}（來源第{item['來源列']}列）"

    return items


def parse_tea(ws):
    items = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = clean(r[1])
        if not name:
            continue
        qty = r[7] if isinstance(r[7], (int, float)) else 0
        unit = clean(r[8])
        cost_per_catty = r[10] if isinstance(r[10], (int, float)) else None
        subtotal = r[9] if isinstance(r[9], (int, float)) else None
        if subtotal is None and cost_per_catty and unit in TAEL_TO_CATTY:
            subtotal = qty * TAEL_TO_CATTY[unit] * cost_per_catty
        items.append(
            {
                "品名": name,
                "品種": clean(r[2]),
                "產地": clean(r[3]),
                "海拔": clean(r[4]),
                "製程": clean(r[5]),
                "庫存數量": qty,
                "單位": unit,
                "庫存成本": subtotal,
                "進貨價_斤": cost_per_catty,
                "產品定價": r[11] if isinstance(r[11], (int, float)) else None,
                "零售價": r[12] if isinstance(r[12], (int, float)) else None,
            }
        )
    return items


def parse(path=SRC, wine_path=WINE_SRC):
    wb = openpyxl.load_workbook(path, data_only=True)
    wine_wb = openpyxl.load_workbook(wine_path, data_only=True)
    return {
        "鑫海產": parse_seafood(wb["鑫海產"]),
        "鑫酒藏": parse_wine(wine_wb["總表"]),
        "鑫茶坊": parse_tea(wb["鑫茶坊"]),
    }


if __name__ == "__main__":
    data = parse(sys.argv[1] if len(sys.argv) > 1 else SRC)
    print(json.dumps(data, ensure_ascii=False, indent=2))
