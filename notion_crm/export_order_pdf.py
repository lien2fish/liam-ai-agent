#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
填入出貨單並輸出指定分頁為 PDF
用法：python3 export_order_pdf.py
"""
import openpyxl
import subprocess
import os

TEMPLATE  = "/Users/lien/Desktop/鉅鑫管理顧問/鑫酒藏/鑫酒藏品項＆客戶紀錄單.xlsx"
TEMP_XLSX = "/tmp/鑫酒藏出貨單_填寫版.xlsx"
PDF_OUT   = "/Users/lien/Desktop/鑫酒藏出貨單_肪煎商行_20260430.pdf"

ITEM_START = 8   # 與 build_xinyou_order.py 一致

ORDER = {
    "訂單編號": "2026043001",
    "訂單日期": "2026/04/30",
    "客戶姓名": "肪煎商行 鄭興安",
    "電話":     "0922-050113",
    "地址":     "新北市新莊區福德二街51號",
    "items": [
        {"品項": "伊拉蘇 Syrah", "類別": "紅酒", "數量": 6, "單位": "瓶", "單價": 570},
    ],
    "折扣": None,
    "運費": None,
    "備註": "",
}

def fill_order():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["出貨單"]

    # ── 標頭 ──
    ws["B3"] = ORDER["訂單編號"]
    ws["F3"] = ORDER["訂單日期"]

    # ── 客戶資訊（覆寫 VLOOKUP 公式為實際值）──
    ws["B4"] = ORDER["客戶姓名"]
    ws["F4"] = ORDER["電話"]
    ws["B5"] = ORDER["地址"]

    # ── 品項明細 ──
    for i, item in enumerate(ORDER["items"]):
        r = ITEM_START + i
        ws.cell(r, 2).value = item["品項"]   # 品項名稱
        ws.cell(r, 3).value = item["類別"]   # 類別（非目錄品項需手動填）
        ws.cell(r, 4).value = item["數量"]   # 數量
        ws.cell(r, 5).value = item["單位"]   # 單位
        ws.cell(r, 6).value = item["單價"]   # 單價（覆寫 VLOOKUP 公式）
        # G 欄公式 =D*F 會自動計算金額

    # ── 折扣、運費、備註 ──
    disc_r  = ITEM_START + 12 + 1  # = 21，對應 build 腳本的 disc_r
    ship_r  = disc_r + 1           # = 22
    note_r  = ship_r + 2           # = 24

    if ORDER["折扣"]:
        ws[f"C{disc_r}"] = ORDER["折扣"]
    if ORDER["運費"]:
        ws[f"G{ship_r}"] = ORDER["運費"]
    if ORDER["備註"]:
        ws[f"B{note_r}"] = ORDER["備註"]

    wb.save(TEMP_XLSX)
    print(f"✓ 暫存填寫版：{TEMP_XLSX}")

def export_pdf():
    script = f'''
tell application "Numbers"
    open POSIX file "{TEMP_XLSX}"
    delay 4
    set doc to front document
    repeat while (count of sheets of doc) > 1
        delete last sheet of doc
    end repeat
    delay 1
    export doc to POSIX file "{PDF_OUT}" as PDF
    close doc saving no
end tell
'''
    result = subprocess.run(["osascript", "-e", script],
                            capture_output=True, text=True)
    if result.returncode == 0:
        print(f"✓ PDF 已儲存：{PDF_OUT}")
    else:
        print(f"✗ AppleScript 錯誤：{result.stderr}")

if __name__ == "__main__":
    fill_order()
    export_pdf()
