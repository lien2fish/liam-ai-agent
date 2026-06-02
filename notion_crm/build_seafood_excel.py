"""
鑫海產 — 品項 & 客戶紀錄單建置腳本
工作表：客戶名單 / 銷售報表 / 出貨單 / 統計儀表板
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ── 色彩常數（深海藍主題）─────────────────────────────────────────────────────
C_NAVY = "0A2342"  # 深海藍（主標題）
C_TEAL = "1B6E8A"  # 青藍（欄標題）
C_TEAL2 = "2A8FAE"  # 淡青藍（次要標題）
C_LIGHT = "D6EEF3"  # 淺海藍（交替列）
C_BG = "F2FAFB"  # 海泡背景
C_WHITE = "FFFFFF"
C_BLACK = "1A1A1A"
C_GREEN = "1A6B3A"  # 毛利正值
C_RED = "B84444"  # 毛利負值
C_GRAY = "F2F2F2"
C_MIDGRAY = "D9D9D9"
C_GOLD = "D4A843"  # 合計列強調


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(name="微軟正黑體", size=10, bold=False, color="1A1A1A", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def border_thin():
    s = Side(style="thin", color=C_MIDGRAY)
    return Border(left=s, right=s, top=s, bottom=s)


def border_medium():
    s = Side(style="medium", color=C_TEAL)
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_mid():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── 客戶預設資料 ──────────────────────────────────────────────────────────────
CUSTOMERS = [
    (
        "HSP-001",
        "吳沛瑾（維甯媽媽）",
        "",
        "LINE",
        "",
        "烏魚子、干貝、透抽",
        "一般",
        6630,
        "成本5800",
    ),
    (
        "HSP-002",
        "張佑全",
        "",
        "LINE",
        "",
        "野生鮑魚、白蝦、透抽、白鯧、龍蝦",
        "VIP",
        10400,
        "成本4060",
    ),
    ("HSP-003", "吳沛瑾", "", "LINE", "", "透抽、黃魚、蛤蜊", "一般", 1600, "成本1000"),
]

# ── 銷售預設資料 ──────────────────────────────────────────────────────────────
SALES = [
    (
        "2026-05-05",
        "HSP-001",
        "吳沛瑾（維甯媽媽）",
        "溏心烏魚子禮盒×1／干貝S×2盒／透抽×3",
        6630,
        5800,
        "LINE轉帳",
        "",
    ),
    (
        "2026-05-05",
        "HSP-002",
        "張佑全",
        "野生鮑魚×12／白蝦×1盒／砲管透抽×2／白鯧1斤×1／迦納1斤×1／干貝M×1盒／龍蝦×2＋加購干貝M×1盒／龍蝦×2",
        10400,
        4060,
        "LINE轉帳",
        "多筆合計",
    ),
    (
        "2026-05-28",
        "HSP-003",
        "吳沛瑾",
        "透抽×3隻、黃魚×2隻、蛤蜊×1斤",
        1600,
        1000,
        "LINE轉帳",
        "",
    ),
]

# ── 建立 Workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ============================================================
# SHEET 1：客戶名單
# ============================================================
ws_cust = wb.active
ws_cust.title = "👥 客戶名單"
ws_cust.sheet_view.showGridLines = False
ws_cust.freeze_panes = "A4"

# 標題
ws_cust.merge_cells("A1:I1")
h = ws_cust["A1"]
h.value = "👥 龜吼現流活海產 — 客戶名單"
h.fill = fill(C_NAVY)
h.font = Font(name="微軟正黑體", size=13, bold=True, color=C_WHITE)
h.alignment = center()
ws_cust.row_dimensions[1].height = 34

# 副標
ws_cust.merge_cells("A2:I2")
sub = ws_cust["A2"]
sub.value = f"建立日期：{date.today()}　　編號規則：HSP-XXX"
sub.fill = fill(C_LIGHT)
sub.font = Font(name="微軟正黑體", size=9, color="555555")
sub.alignment = left_mid()
ws_cust.row_dimensions[2].height = 18

# 欄標題
cust_headers = [
    "客戶編號",
    "姓名",
    "電話",
    "聯絡方式",
    "Email",
    "偏好品項",
    "會員等級",
    "累計消費",
    "備註",
]
cust_widths = [12, 16, 16, 12, 22, 28, 10, 14, 28]
for i, (h_txt, w) in enumerate(zip(cust_headers, cust_widths), 1):
    c = ws_cust.cell(row=3, column=i, value=h_txt)
    c.fill = fill(C_TEAL)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()
    ws_cust.column_dimensions[get_column_letter(i)].width = w
ws_cust.row_dimensions[3].height = 22

# 預填客戶
for ri, (cid, name, phone, contact, email, pref, level, total, note) in enumerate(
    CUSTOMERS, 4
):
    alt = C_BG if ri % 2 == 0 else C_WHITE
    row_data = [cid, name, phone, contact, email, pref, level, total, note]
    for ci, val in enumerate(row_data, 1):
        c = ws_cust.cell(row=ri, column=ci, value=val)
        c.fill = fill(alt)
        c.border = border_thin()
        c.font = font(size=9)
        c.alignment = center() if ci in (1, 4, 7, 8) else left_mid()
    ws_cust.cell(row=ri, column=8).number_format = "#,##0"
    ws_cust.row_dimensions[ri].height = 18

# 空白客戶列（共100列）
for ri in range(4 + len(CUSTOMERS), 104):
    alt = C_BG if ri % 2 == 0 else C_WHITE
    for ci in range(1, 10):
        c = ws_cust.cell(row=ri, column=ci)
        c.fill = fill(alt)
        c.border = border_thin()
        c.font = font(size=9)
    ws_cust.cell(row=ri, column=8).number_format = "#,##0"
    ws_cust.row_dimensions[ri].height = 18

# ============================================================
# SHEET 2：銷售報表
# ============================================================
ws_sales = wb.create_sheet("📊 銷售報表")
ws_sales.sheet_view.showGridLines = False
ws_sales.freeze_panes = "A4"

# 標題
ws_sales.merge_cells("A1:J1")
h = ws_sales["A1"]
h.value = "📊 龜吼現流活海產 — 銷售報表"
h.fill = fill(C_NAVY)
h.font = Font(name="微軟正黑體", size=13, bold=True, color=C_WHITE)
h.alignment = center()
ws_sales.row_dimensions[1].height = 34

# 篩選條件列
ws_sales["A2"].value = "月份："
ws_sales["B2"].value = str(date.today())[:7]
ws_sales["D2"].value = "客戶："
ws_sales["E2"].value = ""
ws_sales["G2"].value = "付款："
ws_sales["H2"].value = ""
for cell_ref in ["A2", "D2", "G2"]:
    ws_sales[cell_ref].font = Font(name="微軟正黑體", size=9, bold=True)
ws_sales.row_dimensions[2].height = 20

# 欄標題
sales_headers = [
    "出貨日期",
    "訂單編號",
    "客戶名稱",
    "品項明細",
    "訂購金額",
    "成本",
    "毛利",
    "毛利率%",
    "付款方式",
    "備註",
]
sales_widths = [12, 12, 18, 50, 12, 12, 12, 10, 14, 22]
for i, (h_txt, w) in enumerate(zip(sales_headers, sales_widths), 1):
    c = ws_sales.cell(row=3, column=i, value=h_txt)
    c.fill = fill(C_TEAL)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()
    ws_sales.column_dimensions[get_column_letter(i)].width = w
ws_sales.row_dimensions[3].height = 22

# 預填銷售資料
for ri, (dt, oid, name, items, amount, cost, pay, note) in enumerate(SALES, 4):
    alt = C_BG if ri % 2 == 0 else C_WHITE
    row_data = [dt, oid, name, items, amount, cost, None, None, pay, note]
    for ci, val in enumerate(row_data, 1):
        c = ws_sales.cell(row=ri, column=ci, value=val)
        c.fill = fill(alt)
        c.border = border_thin()
        c.font = font(size=9)
        c.alignment = center() if ci in (1, 2, 5, 6, 7, 8, 9) else left_mid()
    # 毛利公式
    ws_sales.cell(row=ri, column=7).value = f"=E{ri}-F{ri}"
    ws_sales.cell(row=ri, column=7).number_format = "#,##0"
    # 毛利率公式
    ws_sales.cell(row=ri, column=8).value = f'=IFERROR(G{ri}/E{ri},"")'
    ws_sales.cell(row=ri, column=8).number_format = "0.0%"
    # 金額格式
    for ci in (5, 6):
        ws_sales.cell(row=ri, column=ci).number_format = "#,##0"
    ws_sales.row_dimensions[ri].height = 18

# 空白銷售列（共500列）
start_r = 4 + len(SALES)
for r in range(start_r, 504):
    alt = C_BG if r % 2 == 0 else C_WHITE
    for col in range(1, 11):
        c = ws_sales.cell(row=r, column=col)
        c.fill = fill(alt)
        c.border = border_thin()
        c.font = font(size=9)
        c.alignment = center() if col in (1, 2, 5, 6, 7, 8, 9) else left_mid()
    ws_sales.cell(row=r, column=7).value = f'=IF(E{r}="","",E{r}-F{r})'
    ws_sales.cell(row=r, column=7).number_format = "#,##0"
    ws_sales.cell(row=r, column=8).value = f'=IFERROR(G{r}/E{r},"")'
    ws_sales.cell(row=r, column=8).number_format = "0.0%"
    for ci in (5, 6):
        ws_sales.cell(row=r, column=ci).number_format = "#,##0"
    ws_sales.row_dimensions[r].height = 17

# 合計列
sum_row = 504
ws_sales.merge_cells(f"A{sum_row}:D{sum_row}")
ws_sales[f"A{sum_row}"].value = "▶ 期間合計"
ws_sales[f"A{sum_row}"].font = Font(name="微軟正黑體", size=10, bold=True, color=C_NAVY)
ws_sales[f"A{sum_row}"].alignment = Alignment(horizontal="right")
ws_sales[f"A{sum_row}"].fill = fill(C_LIGHT)
for col_idx, col_letter in [(5, "E"), (6, "F"), (7, "G")]:
    c = ws_sales.cell(row=sum_row, column=col_idx)
    c.value = f"=SUM({col_letter}4:{col_letter}503)"
    c.number_format = "#,##0"
    c.font = Font(name="微軟正黑體", size=11, bold=True, color=C_NAVY)
    c.fill = fill(C_LIGHT)
    c.alignment = Alignment(horizontal="right")
    c.border = border_medium()
ws_sales.row_dimensions[sum_row].height = 26

# ============================================================
# SHEET 3：出貨單
# ============================================================
ws_ship = wb.create_sheet("📦 出貨單")
ws_ship.sheet_view.showGridLines = False

# 設定欄寬
for col, w in [("A", 14), ("B", 22), ("C", 14), ("D", 22), ("E", 14), ("F", 18)]:
    ws_ship.column_dimensions[col].width = w

ws_ship.merge_cells("A1:F1")
h = ws_ship["A1"]
h.value = "📦 龜吼現流活海產 — 出貨單"
h.fill = fill(C_NAVY)
h.font = Font(name="微軟正黑體", size=14, bold=True, color=C_WHITE)
h.alignment = center()
ws_ship.row_dimensions[1].height = 38

# 基本資訊欄位
fields = [
    ("A2", "出貨單號"),
    ("B2", ""),
    ("D2", "出貨日期"),
    ("E2", str(date.today())),
    ("A3", "客戶名稱"),
    ("B3", ""),
    ("D3", "電話"),
    ("E3", ""),
    ("A4", "付款方式"),
    ("B4", ""),
    ("D4", "業務"),
    ("E4", ""),
]
for cell_ref, val in fields:
    c = ws_ship[cell_ref]
    c.value = val
    if val and not val.startswith("2026") and val not in ("",):
        try:
            ws_ship[cell_ref]
        except:
            pass
for cell_ref in ["A2", "D2", "A3", "D3", "A4", "D4"]:
    ws_ship[cell_ref].font = Font(name="微軟正黑體", size=9, bold=True, color=C_NAVY)
    ws_ship[cell_ref].fill = fill(C_LIGHT)
    ws_ship[cell_ref].alignment = center()
    ws_ship[cell_ref].border = border_thin()
for cell_ref in ["B2", "E2", "B3", "E3", "B4", "E4"]:
    ws_ship[cell_ref].border = border_thin()
    ws_ship[cell_ref].font = font(size=10)
for r in [2, 3, 4]:
    ws_ship.row_dimensions[r].height = 22

# 品項標題
ws_ship.row_dimensions[6].height = 22
item_headers = ["品項名稱", "規格/數量", "單位", "單價", "小計", "備註"]
item_widths = [22, 16, 10, 12, 14, 20]
for i, (h_txt, w) in enumerate(zip(item_headers, item_widths), 1):
    c = ws_ship.cell(row=6, column=i, value=h_txt)
    c.fill = fill(C_TEAL)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()

# 品項列（25列）
for r in range(7, 32):
    alt = C_BG if r % 2 == 0 else C_WHITE
    for ci in range(1, 7):
        c = ws_ship.cell(row=r, column=ci)
        c.fill = fill(alt)
        c.border = border_thin()
        c.font = font(size=9)
        c.alignment = center() if ci in (3, 4, 5) else left_mid()
    ws_ship.cell(row=r, column=5).value = f'=IF(D{r}="","",D{r}*C{r})'
    ws_ship.cell(row=r, column=4).number_format = "#,##0"
    ws_ship.cell(row=r, column=5).number_format = "#,##0"
    ws_ship.row_dimensions[r].height = 18

# 合計
ws_ship.merge_cells("A32:D32")
ws_ship["A32"].value = "合計金額"
ws_ship["A32"].font = Font(name="微軟正黑體", size=10, bold=True, color=C_WHITE)
ws_ship["A32"].fill = fill(C_TEAL2)
ws_ship["A32"].alignment = Alignment(horizontal="right")
ws_ship["A32"].border = border_thin()
ws_ship["E32"].value = "=SUM(E7:E31)"
ws_ship["E32"].number_format = "#,##0"
ws_ship["E32"].font = Font(name="微軟正黑體", size=12, bold=True, color=C_NAVY)
ws_ship["E32"].fill = fill(C_LIGHT)
ws_ship["E32"].alignment = center()
ws_ship["E32"].border = border_medium()
ws_ship.row_dimensions[32].height = 26

# 簽收
ws_ship.merge_cells("A34:C34")
ws_ship["A34"].value = "收貨簽收：___________________________"
ws_ship["A34"].font = font(size=10)
ws_ship.merge_cells("D34:F34")
ws_ship["D34"].value = "日期：_______________"
ws_ship["D34"].font = font(size=10)
ws_ship.row_dimensions[34].height = 24

# ============================================================
# SHEET 4：統計儀表板
# ============================================================
ws_dash = wb.create_sheet("📈 統計儀表板")
ws_dash.sheet_view.showGridLines = False

ws_dash.merge_cells("A1:L1")
h = ws_dash["A1"]
h.value = "📈 龜吼現流活海產 — 銷售統計儀表板"
h.fill = fill(C_NAVY)
h.font = Font(name="微軟正黑體", size=14, bold=True, color=C_WHITE)
h.alignment = center()
ws_dash.row_dimensions[1].height = 36

# KPI 標題
ws_dash.merge_cells("A3:L3")
ws_dash["A3"].value = "月度業績 KPI"
ws_dash["A3"].font = Font(name="微軟正黑體", size=11, bold=True, color=C_TEAL)
ws_dash["A3"].fill = fill(C_LIGHT)
ws_dash["A3"].alignment = left_mid()

kpis = [
    (
        "本月銷售總額",
        f"=SUMIF('📊 銷售報表'!A$4:A$503,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'📊 銷售報表'!E$4:E$503)",
        C_NAVY,
    ),
    (
        "本月出貨筆數",
        f"=COUNTIF('📊 銷售報表'!A$4:A$503,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1))",
        C_TEAL,
    ),
    (
        "本月毛利合計",
        f"=SUMIF('📊 銷售報表'!A$4:A$503,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'📊 銷售報表'!G$4:G$503)",
        C_GREEN,
    ),
    ("最高客單", "=IFERROR(MAX('📊 銷售報表'!E$4:E$503),0)", C_TEAL2),
    (
        "平均客單",
        "=IFERROR(AVERAGEIF('📊 銷售報表'!E$4:E$503,\">0\",'📊 銷售報表'!E$4:E$503),0)",
        C_BLACK,
    ),
    ("累計銷售總額", "=SUM('📊 銷售報表'!E$4:E$503)", C_GOLD),
]

for ki, (label, formula, color) in enumerate(kpis):
    col = 1 + (ki % 3) * 4
    row_start = 5 + (ki // 3) * 5
    ws_dash.merge_cells(
        start_row=row_start, start_column=col, end_row=row_start, end_column=col + 3
    )
    ws_dash.merge_cells(
        start_row=row_start + 1,
        start_column=col,
        end_row=row_start + 1,
        end_column=col + 3,
    )
    lc = ws_dash.cell(row=row_start, column=col, value=label)
    lc.font = Font(name="微軟正黑體", size=8, color="777777")
    lc.alignment = center()
    lc.fill = fill(C_GRAY)
    vc = ws_dash.cell(row=row_start + 1, column=col, value=formula)
    vc.font = Font(name="微軟正黑體", size=20, bold=True, color=color)
    vc.alignment = center()
    vc.fill = fill(C_BG)
    vc.number_format = "#,##0"
    for c_idx in range(col, col + 4):
        ws_dash.cell(row=row_start, column=c_idx).border = border_thin()
        ws_dash.cell(row=row_start + 1, column=c_idx).border = border_thin()
    ws_dash.row_dimensions[row_start].height = 16
    ws_dash.row_dimensions[row_start + 1].height = 34

# 客戶排行
ws_dash["A17"].value = "客戶排行（手動維護）"
ws_dash["A17"].font = Font(name="微軟正黑體", size=11, bold=True, color=C_TEAL)
ws_dash["A17"].fill = fill(C_LIGHT)
ws_dash.merge_cells("A17:D17")

rank_headers = ["排名", "客戶名稱", "累計消費(NT$)", "訂單數"]
for ci, h_txt in enumerate(rank_headers, 1):
    c = ws_dash.cell(row=18, column=ci, value=h_txt)
    c.fill = fill(C_TEAL)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()
    ws_dash.column_dimensions[get_column_letter(ci)].width = [6, 20, 16, 10][ci - 1]

rank_data = [
    (1, "張佑全", 10400, 1),
    (2, "吳沛瑾（維甯媽媽）", 6630, 1),
    (3, "吳沛瑾", 1600, 1),
]
for ri, (rank, name, total, cnt) in enumerate(rank_data, 19):
    alt = C_BG if ri % 2 == 0 else C_WHITE
    for ci, val in enumerate([rank, name, total, cnt], 1):
        c = ws_dash.cell(row=ri, column=ci, value=val)
        c.fill = fill(alt)
        c.border = border_thin()
        c.font = font(size=9)
        c.alignment = center() if ci in (1, 3, 4) else left_mid()
    ws_dash.cell(row=ri, column=3).number_format = "#,##0"
    ws_dash.row_dimensions[ri].height = 18

# 使用說明
tips_start = 24
ws_dash.merge_cells(f"A{tips_start}:L{tips_start}")
ws_dash[f"A{tips_start}"].value = "💡 使用說明"
ws_dash[f"A{tips_start}"].font = Font(
    name="微軟正黑體", size=10, bold=True, color=C_TEAL
)
ws_dash[f"A{tips_start}"].fill = fill(C_LIGHT)
ws_dash.merge_cells(f"A{tips_start}:L{tips_start}")

tips = [
    "1. 【客戶名單】：新客戶照 HSP-XXX 流水號建立，偏好品項請即時更新",
    "2. 【銷售報表】：每筆訂單填入出貨日期/訂單編號/品項明細/金額/成本，毛利與毛利率自動計算",
    "3. 【出貨單】：每次出貨列印此頁，客戶簽收後留存",
    "4. 【統計儀表板】：KPI 自動從銷售報表彙整；客戶排行請手動更新",
    "5. 訂單編號與 Notion CRM 同步，Notion 為主要查詢介面",
]
for ti, tip in enumerate(tips):
    r = tips_start + 1 + ti
    ws_dash.merge_cells(f"A{r}:L{r}")
    c = ws_dash.cell(row=r, column=1, value=tip)
    c.font = Font(name="微軟正黑體", size=9)
    c.alignment = left_mid()
    c.fill = fill(C_BG if ti % 2 == 0 else C_WHITE)
    ws_dash.row_dimensions[r].height = 17

# ── 列印設定 ─────────────────────────────────────────────────────────────────
for ws in [ws_cust, ws_sales, ws_ship]:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

# ── Tab 顏色 ────────────────────────────────────────────────────────────────
ws_cust.sheet_properties.tabColor = "1B6E8A"
ws_sales.sheet_properties.tabColor = "2A8FAE"
ws_ship.sheet_properties.tabColor = "1A6B3A"
ws_dash.sheet_properties.tabColor = "0A2342"

# ── 儲存 ────────────────────────────────────────────────────────────────────
output_path = "/Users/lien/Desktop/鉅鑫管理顧問/鑫海產/鑫海產品項＆客戶紀錄單.xlsx"
wb.save(output_path)
print(f"✅ 已儲存：{output_path}")
print(f"   客戶名單：{len(CUSTOMERS)} 筆預填")
print(f"   銷售報表：{len(SALES)} 筆預填，500 列可用")
print(f"   出貨單：25 列可填入")
print(f"   統計儀表板：自動彙整 KPI")
