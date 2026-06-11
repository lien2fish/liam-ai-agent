"""
星坊酒業 SERGIO VALENTE — 販售管理 Excel 製作腳本
包含：商品目錄、出貨單、銷售報表
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from datetime import date
import re

# ── 色彩常數 ──────────────────────────────────────────────────────────────────
C_DARKGOLD = "8B4513"  # 深棕
C_GOLD = "B7874A"  # 金棕
C_LIGHTGOLD = "F0E6D3"  # 淡奶油
C_CREAM = "FAF8F3"  # 背景奶油
C_WHITE = "FFFFFF"
C_BLACK = "1A1A1A"
C_GREEN = "3D7A52"
C_RED = "B84444"
C_GRAY = "F2F2F2"
C_MIDGRAY = "D9D9D9"
C_HEADER = "2D1B0E"  # 深棕標題列


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(name="微軟正黑體", size=10, bold=False, color="1A1A1A", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def border_thin():
    s = Side(style="thin", color=C_MIDGRAY)
    return Border(left=s, right=s, top=s, bottom=s)


def border_medium():
    s = Side(style="medium", color=C_GOLD)
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_middle():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── 商品資料 ───────────────────────────────────────────────────────────────────
PRODUCTS = [
    # ── 紅酒 (90 筆) ──
    (
        "ARM-101J211A",
        "To Kalon Reserve Cabernet Sauvignon 2021",
        "ROBERT MONDAVI WINERY",
        "紅酒",
        9700,
    ),
    (
        "ARM-301J221A",
        "Oakville Cabernet Sauvignon '22",
        "ROBERT MONDAVI WINERY",
        "紅酒",
        4250,
    ),
    ("ARM-302J211A", "Carneros Pinot Noir '21", "ROBERT MONDAVI WINERY", "紅酒", 4700),
    (
        "ARM-405J211A",
        "Spotlight Cabernet Sauvignon '21",
        "ROBERT MONDAVI WINERY",
        "紅酒",
        3350,
    ),
    (
        "ARM-401J211A",
        "Napa Valley Cabernet Sauvignon '21",
        "ROBERT MONDAVI WINERY",
        "紅酒",
        2850,
    ),
    (
        "ARM-514J241A",
        "California Cabernet Sauvignon 2024",
        "ROBERT MONDAVI WINERY",
        "紅酒",
        1200,
    ),
    (
        "ARM-515J241A",
        "California Pinot Noir 2024",
        "ROBERT MONDAVI WINERY",
        "紅酒",
        1200,
    ),
    ("ARM-701J", "Woodbridge Cabernet Sauvignon", "ROBERT MONDAVI WINERY", "紅酒", 760),
    ("ACT-101J211A", "CONTINUUM '21", "CONTINUUM ESTATE", "紅酒", 19400),
    ("ACT-301J191A", "初心紅葡萄酒 2019", "CONTINUUM ESTATE", "紅酒", 10950),
    ("ARW-102J201A", "西岸自由之石 博德加 黑皮諾 2020", "RAEN", "紅酒", 6950),
    ("ARW-201J231A", "索諾瑪海岸 聖羅伯 黑皮諾 2023", "RAEN", "紅酒", 4900),
    ("AWF-101J191A", "特選 卡本內蘇維濃 2019", "CAYMUS VINEYARDS", "紅酒", 10800),
    ("AWF-301J231A", "那帕山谷 卡本內蘇維濃 2023", "CAYMUS VINEYARDS", "紅酒", 4350),
    ("AWI-501J211A", "紅酒 2021", "CONUNDRUM", "紅酒", 1980),
    ("AWJ-102J201A", "那帕山谷 梅洛紅酒 2020", "EMMOLO", "紅酒", 3780),
    ("AWG-102J211A", "酪農園 黑皮諾 2021", "BELLE GLOS", "紅酒", 3220),
    ("AWG-103J221A", "高地園 黑皮諾 2022", "BELLE GLOS", "紅酒", 3220),
    ("AWG-104J231A", "Clark & Telephone Pinot Noir '23", "BELLE GLOS", "紅酒", 3220),
    ("ASO-201J191A", "Napa Valley Cabernet Sauvignon '19", "SILVER OAK", "紅酒", 11500),
    (
        "ASO-301J211A",
        "Alexander Valley Cabernet Sauvignon '21",
        "SILVER OAK",
        "紅酒",
        6000,
    ),
    ("ATL-101J201A", "Timeless Napa Valley Red Wine '20", "TIMELESS", "紅酒", 13000),
    ("ASP-202J191A", "Bien Nacido Pinot Noir '19", "TWOMEY", "紅酒", 3760),
    ("ASP-302J211A", "Russian River Valley Pinot Noir '21", "TWOMEY", "紅酒", 4200),
    ("ASP-304J221A", "Anderson Valley Pinot Noir '22", "TWOMEY", "紅酒", 4200),
    ("AOD-101J211A", "OVID Napa Valley Red Wine '21", "OVID", "紅酒", 15000),
    ("AOD-102J211A", "OVID Hexameter '21", "OVID", "紅酒", 15000),
    ("ACV-101J211A", "Yettalil Napa Valley '21", "CLOS DU VAL", "紅酒", 12500),
    ("ACV-401J221A", "Napa Valley Cabernet Sauvignon '22", "CLOS DU VAL", "紅酒", 3580),
    ("ACV-501J211A", "Red Blend '21", "CLOS DU VAL", "紅酒", 2250),
    ("ADE-101J211A", "Dominus '21", "DOMINUS ESTATE", "紅酒", 26000),
    ("ADE-401J201A", "Napanook '20", "DOMINUS ESTATE", "紅酒", 7030),
    ("ADE-501J191A", "Othello '19", "DOMINUS ESTATE", "紅酒", 3700),
    ("AUV-401J191A", "Ulysses '19", "ULYSSES VINEYARD", "紅酒", 13000),
    (
        "AJV-301J191A",
        "Alexander Valley Cabernet Sauvignon '19",
        "JORDAN VINEYARD & WINERY",
        "紅酒",
        4100,
    ),
    (
        "ACL-302J211A",
        "Archimedes Cabernet Sauvignon '21",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        4250,
    ),
    (
        "ACL-401J201A",
        "Eleanor Red Wine '20",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        3300,
    ),
    (
        "ACL-301J211A",
        "Director's Cut Alexander Valley Cabernet Sauvignon '21",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        2250,
    ),
    (
        "ACL-402J221A",
        "Director's Cut Pinot Noir '22",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        2150,
    ),
    (
        "ACL-501J231A",
        "Diamond Claret Cabernet Sauvignon '23",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        1300,
    ),
    (
        "ACL-502J221A",
        "Diamond Pinot Noir '22",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        1200,
    ),
    ("ACL-503J211A", "Diamond Merlot '21", "FRANCIS FORD COPPOLA WINERY", "紅酒", 1200),
    (
        "ACL-504J211A",
        "Diamond Zinfandel '21",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        1200,
    ),
    ("ACL-505J201A", "Diamond Red '20", "FRANCIS FORD COPPOLA WINERY", "紅酒", 1380),
    (
        "ACL-701J231A",
        "Rosso & Bianco Cabernet Sauvignon '23",
        "FRANCIS FORD COPPOLA WINERY",
        "紅酒",
        760,
    ),
    ("AKD-101J181A", "Cardinale Cabernet Sauvignon '18", "CARDINALE", "紅酒", 19800),
    ("AEZ-104J121A", "Kai Carmenere '12", "VIÑA ERRÁZURIZ", "紅酒", 8000),
    (
        "AAV-101J211A",
        "Sleeping Lady Vineyard Cabernet Sauvignon '21",
        "AD VIVUM",
        "紅酒",
        8800,
    ),
    ("AEZ-103J121A", "La Cumbre Syrah '12", "VIÑA ERRÁZURIZ", "紅酒", 4670),
    ("ABO-101J171A", "Proprietary Red Wine '17", "BELLA OAKS VINEYARD", "紅酒", 14000),
    (
        "ABO-401J211A",
        "Le Génie Cabernet Sauvignon 2021",
        "BELLA OAKS VINEYARD",
        "紅酒",
        7000,
    ),
    (
        "ARC-102J221A",
        "Beckstoffer Dr. Crane Vineyard Cabernet Sauvignon '22",
        "REALM CELLARS",
        "紅酒",
        26800,
    ),
    ("ARC-301J221A", "Houyi Cabernet Sauvignon '22", "REALM CELLARS", "紅酒", 15000),
    (
        "ARC-303J221A",
        "Moonracer Cabernet Sauvignon '22",
        "REALM CELLARS",
        "紅酒",
        15000,
    ),
    (
        "ARC-304J221A",
        "Hartwell XX Cabernet Sauvignon '22",
        "REALM CELLARS",
        "紅酒",
        28000,
    ),
    ("ARC-401J221A", "The Bard Cabernet Sauvignon '22", "REALM CELLARS", "紅酒", 6800),
    ("ARC-402J221A", "The Tempest '22", "REALM CELLARS", "紅酒", 6800),
    (
        "AVI-101J211A",
        "Beckstoffer Dr. Crane Vineyard Cabernet Sauvignon '21",
        "VICE VERSA WINES",
        "紅酒",
        14500,
    ),
    (
        "AVI-103J211A",
        "Beckstoffer Las Piedras Vineyard Cabernet Sauvignon '21",
        "VICE VERSA WINES",
        "紅酒",
        14500,
    ),
    (
        "AVI-104J211A",
        "Beckstoffer Georges III Vineyard Cabernet Sauvignon '21",
        "VICE VERSA WINES",
        "紅酒",
        14500,
    ),
    (
        "AVI-402J211A",
        "Mysterons Cabernet Sauvignon '21",
        "VICE VERSA WINES",
        "紅酒",
        15500,
    ),
    (
        "AVI-403J211A",
        "Spinning Plates Cabernet Sauvignon '21",
        "VICE VERSA WINES",
        "紅酒",
        6500,
    ),
    ("AST-101J191A", "STONES No. 1 '19", "STONE WINE", "紅酒", 16800),
    ("AST-102J191A", "STONES No. 2 '19", "STONE WINE", "紅酒", 16800),
    ("AST-103J191A", "STONES No. 3 '19", "STONE WINE", "紅酒", 16800),
    ("AST-301J191A", "Longitude-Latitude '19", "STONE WINE", "紅酒", 16800),
    (
        "AST-001J211A",
        "TICKETS by Stones Cabernet Sauvignon '21",
        "STONE WINE",
        "紅酒",
        46800,
    ),
    (
        "AFJ-501J221A",
        "East Side Ridge Cabernet Sauvignon '22",
        "Three Finger Jack",
        "紅酒",
        1100,
    ),
    ("AFJ-502J211A", "Old Vine Zinfandel '21", "Three Finger Jack", "紅酒", 1100),
    (
        "AFJ-503J211A",
        "Rum Barrel Aged Red Blend '21",
        "Three Finger Jack",
        "紅酒",
        1100,
    ),
    ("AYE-102J171A", "Ironheart Shiraz '17", "YANGARRA ESTATE VINEYARD", "紅酒", 4560),
    ("AYE-702J191A", "Estate Noir '19", "YANGARRA ESTATE VINEYARD", "紅酒", 1410),
    ("AMD-104J221A", "Velvet Glove '22", "MOLLYDOOKER WINERY", "紅酒", 8850),
    ("AJS-704J201A", "Chateau Plince '20", "ÉTS. JEAN-PIERRE MOUEIX", "紅酒", 3100),
    ("AJM-702J161A", "JPM Medoc '16", "ÉTS. JEAN-PIERRE MOUEIX", "紅酒", 990),
    ("AJM-701J161A", "JPM Bordeaux '16", "ÉTS. JEAN-PIERRE MOUEIX", "紅酒", 890),
    (
        "AJS-401J201A",
        "Annonce de Belair-Monange '20",
        "ÉTS. JEAN-PIERRE MOUEIX",
        "紅酒",
        4930,
    ),
    ("AJS-701J171A", "Haut Roc Blanquant '17", "ÉTS. JEAN-PIERRE MOUEIX", "紅酒", 2530),
    (
        "AJS-203J171A",
        "Chateau Belair-Monange '17",
        "ÉTS. JEAN-PIERRE MOUEIX",
        "紅酒",
        11050,
    ),
    (
        "AJS-703J191A",
        "Chateau Lafleur-Gazin '19",
        "ÉTS. JEAN-PIERRE MOUEIX",
        "紅酒",
        3130,
    ),
    (
        "AJS-404J141A",
        "Chateau Puy-Blanquet '14",
        "ÉTS. JEAN-PIERRE MOUEIX",
        "紅酒",
        1530,
    ),
    (
        "AJS-705J201A",
        "Chateau Chantalouette '20",
        "ÉTS. JEAN-PIERRE MOUEIX",
        "紅酒",
        2330,
    ),
    ("AJS-304J171A", "Chateau La Serre '17", "ÉTS. JEAN-PIERRE MOUEIX", "紅酒", 3830),
    (
        "AJS-403J211A",
        "Chateau Pindefleurs '21",
        "ÉTS. JEAN-PIERRE MOUEIX",
        "紅酒",
        2130,
    ),
    (
        "AAQ-304J201A",
        "CHATEAU PICQUE CAILLOU Rouge '20",
        "CHATEAU PICQUE CAILLOU",
        "紅酒",
        2400,
    ),
    (
        "ACF-305J211A",
        "BARON de LESTAC Rouge, Signature '21",
        "CASTEL FRÈRES",
        "紅酒",
        780,
    ),
    (
        "ACF-201J211A",
        "ASPIRANT DE BEYCHEVELLE '21",
        "CHÂTEAU BEYCHEVELLE",
        "紅酒",
        2430,
    ),
    ("ACC-101J171A", "Clos de Vougeot Grand Cru '17", "Domaine Chanson", "紅酒", 16300),
    (
        "ACC-102J131A",
        "Charmes Chambertin Grand Cru 2013",
        "Domaine Chanson",
        "紅酒",
        12500,
    ),
    ("ACC-211J221A", "Beaune 1er Cru Le Bastion '22", "Domaine Chanson", "紅酒", 3000),
    # ── 白酒 ──
    (
        "ARM-122J221A",
        "To Kalon Reserve Fume Blanc '22",
        "ROBERT MONDAVI WINERY",
        "白酒",
        3850,
    ),
    (
        "ARM-421J221A",
        "Napa Valley Chardonnay '22",
        "ROBERT MONDAVI WINERY",
        "白酒",
        2100,
    ),
    (
        "ARM-422J221A",
        "Napa Valley Sauvignon Blanc '22",
        "ROBERT MONDAVI WINERY",
        "白酒",
        2100,
    ),
    (
        "ARM-525J241A",
        "California Chardonnay 2024",
        "ROBERT MONDAVI WINERY",
        "白酒",
        1200,
    ),
    ("ARM-721J", "Woodbridge Chardonnay", "ROBERT MONDAVI WINERY", "白酒", 760),
    ("ARM-722J", "Woodbridge Sauvignon Blanc", "ROBERT MONDAVI WINERY", "白酒", 760),
    ("ARW-221J231A", "Sonoma Coast Lady Marjorie Chardonnay '23", "RAEN", "白酒", 4900),
    ("AWI-521J221A", "White '22", "CONUNDRUM", "白酒", 1730),
    ("AWJ-322J231A", "Sauvignon Blanc '23", "EMMOLO", "白酒", 1650),
    ("AWH-121J231A", "Reserve Chardonnay '23", "MER SOLEIL", "白酒", 2150),
    ("AWH-321J231A", "Silver Chardonnay '23", "MER SOLEIL", "白酒", 1700),
    ("ASP-422J231A", "Sauvignon Blanc '23", "TWOMEY", "白酒", 2600),
    ("ACT-421J231A", "SENTIUM Sauvignon Blanc '23", "SENTIUM", "白酒", 4600),
    ("ACV-422J231A", "Napa Valley Sauvignon Blanc", "CLOS DU VAL", "白酒", 3400),
    (
        "AJV-321J221A",
        "Russian River Valley Chardonnay '22",
        "JORDAN VINEYARD & WINERY",
        "白酒",
        2680,
    ),
    (
        "ACL-421J221A",
        "Director's Cut Chardonnay '22",
        "FRANCIS FORD COPPOLA WINERY",
        "白酒",
        1950,
    ),
    (
        "ACL-521J231A",
        "Diamond Chardonnay '23",
        "FRANCIS FORD COPPOLA WINERY",
        "白酒",
        1200,
    ),
    (
        "ACL-522J231A",
        "Diamond Vibrance Pinot Grigio '23",
        "FRANCIS FORD COPPOLA WINERY",
        "白酒",
        1110,
    ),
    ("ACL-721J241A", "R&B Chardonnay 2024", "FRANCIS FORD COPPOLA WINERY", "白酒", 760),
    ("ARC-421J241A", "Federigo Sauvignon Blanc 2024", "REALM CELLARS", "白酒", 4600),
    (
        "AST-321J201A",
        "Russian River Valley Chardonnay 2020",
        "STONE WINE",
        "白酒",
        11000,
    ),
    ("AFJ-521J231A", "Gold Mine Chardonnay 2023", "Three Finger Jack", "白酒", 1100),
    ("ATB-124J181A", "Clos Sainte Hune Riesling 2018", "TRIMBACH", "白酒", 18700),
    (
        "ATB-361J171A",
        "Gewurztraminer Vendanges Tardives 2017",
        "TRIMBACH",
        "白酒",
        4380,
    ),
    ("ATB-121J161A", "Frédéric Emile Riesling 2016", "TRIMBACH", "白酒", 5380),
    (
        "ATB-122J151A",
        "Seigneurs de Ribeaupierre Gewurztraminer 2015",
        "TRIMBACH",
        "白酒",
        3450,
    ),
    ("ATB-522J191A", "Gewurztraminer", "TRIMBACH", "白酒", 1980),
    (
        "AAQ-304J201B",
        "CHATEAU PICQUE CAILLOU Blanc '21",
        "CHATEAU PICQUE CAILLOU",
        "白酒",
        2680,
    ),
    (
        "ACF-305J221B",
        "BARON de LESTAC Blanc, Signature '22",
        "CASTEL FRÈRES",
        "白酒",
        780,
    ),
    ("ACC-301J221A", "Corton Vergennes Grand Cru '22", "Domaine Chanson", "白酒", 8400),
    ("ACC-202J231A", "Chablis 1er Cru Montmains '23", "Domaine Chanson", "白酒", 3050),
    (
        "ACC-203J221A",
        "Pernand Vergelesses Les Caradeux 1er Cru '22",
        "Domaine Chanson",
        "白酒",
        3300,
    ),
    (
        "ACC-204J231A",
        "Chassagne Montrachet Les Chenevottes 1er Cru '23",
        "Domaine Chanson",
        "白酒",
        5850,
    ),
    (
        "ACC-205J221A",
        "Beaune Clos Des Mouches 1er Cru '22",
        "Domaine Chanson",
        "白酒",
        6450,
    ),
    ("ACC-206J221A", "Pernand Vergelesses White '22", "Domaine Chanson", "白酒", 2780),
    ("ACC-207J221A", "Chassagne-Montrachet '22", "Domaine Chanson", "白酒", 4650),
    # ── 粉紅酒 ──
    ("ARM-742JNV1A", "Woodbridge Rosé", "ROBERT MONDAVI WINERY", "粉紅酒", 760),
    ("AMA-942J241A", "Rosa dei Masi '24", "MASI", "粉紅酒", 1360),
    ("AKC-341J221A", "Rose '22", "KIM CRAWFORD", "粉紅酒", 1160),
    # ── 氣泡/香檳 ──
    (
        "ACB-181J161A",
        "BOLLINGER Vieilles Vignes Françaises (VVF) '16",
        "CHAMPAGNE BOLLINGER",
        "氣泡/香檳",
        97700,
    ),
    (
        "ACB-281J081A",
        "Bollinger R.D. (Recently Disgorged) '08",
        "CHAMPAGNE BOLLINGER",
        "氣泡/香檳",
        17600,
    ),
    (
        "ACB-486J201A",
        "BOLLINGER PN Blanc de Noirs (TX 20)",
        "CHAMPAGNE BOLLINGER",
        "氣泡/香檳",
        7000,
    ),
    (
        "ACB-383J141A",
        "BOLLINGER La Grande Année Rosé '14",
        "CHAMPAGNE BOLLINGER",
        "氣泡/香檳",
        12400,
    ),
    (
        "ACB-381J151A",
        "BOLLINGER La Grande Année '15",
        "CHAMPAGNE BOLLINGER",
        "氣泡/香檳",
        9680,
    ),
    ("ACB-385J161A", "BOLLINGER B16", "CHAMPAGNE BOLLINGER", "氣泡/香檳", 8300),
    ("ACB-482JNV1A", "BOLLINGER Rosé NV", "CHAMPAGNE BOLLINGER", "氣泡/香檳", 6300),
    (
        "ACB-481JNV1A",
        "BOLLINGER Special Cuvee NV",
        "CHAMPAGNE BOLLINGER",
        "氣泡/香檳",
        4900,
    ),
    ("ACA-181J151A", "AYALA La Perle '15", "CHAMPAGNE AYALA", "氣泡/香檳", 11000),
    ("ACA-381J181A", "AYALA Blanc de Blancs '18", "CHAMPAGNE AYALA", "氣泡/香檳", 5460),
    ("ACA-482JNV1A", "AYALA Rosé Majeur NV", "CHAMPAGNE AYALA", "氣泡/香檳", 3950),
    (
        "ACA-483JNV1A",
        "AYALA Brut Natural (Zero-dosage) NV",
        "CHAMPAGNE AYALA",
        "氣泡/香檳",
        3480,
    ),
    ("ACA-481JNV1A", "AYALA Brut Majeur NV", "CHAMPAGNE AYALA", "氣泡/香檳", 3250),
    (
        "ACF-382JNV1A",
        "Champagne Malard Premium Brut NV",
        "CHAMPAGNE MALARD",
        "氣泡/香檳",
        2350,
    ),
    (
        "ASW-971JNV1A",
        "Monsieur WILLIAM Peach Cocktail Sparkling NV",
        "MONSIEUR WILLIAM",
        "氣泡/香檳",
        480,
    ),
    ("ASG-171JNV1A", "BRUT RESERVE NV", "SCHLOSS GOBELSBURG", "氣泡/香檳", 2300),
    ("ATA-271JNV2A", "Asti Dolce DOCG NV", "TOSTI", "氣泡/香檳", 790),
    ("ATA-273JNV1A", "Pink Moscato NV", "TOSTI", "氣泡/香檳", 790),
    ("ATA-275JNV1A", "ASTI SECCO DOCG NV", "TOSTI", "氣泡/香檳", 895),
    (
        "AFD-971JNV1A",
        "Danzante Prosecco Spumante Extra Dry NV",
        "DANZANTE",
        "氣泡/香檳",
        1250,
    ),
    ("AMC-171J221A", "Cartizze Superiore Dry '22", "CANEVEL", "氣泡/香檳", 3300),
    ("AMC-172J221A", "Campofalco Brut '22", "CANEVEL", "氣泡/香檳", 1830),
    ("AMC-173J211A", "Il Millesimato Extra Dry '21", "CANEVEL", "氣泡/香檳", 1830),
    ("AMC-271J211A", "Canevel Brut Setàge '21", "CANEVEL", "氣泡/香檳", 1500),
    # ── 清酒 ──
    ("CSF-6151", "Shuho Junmai Daiginjo Yamadaho 22", "秀鳳酒造場", "清酒", 2880),
    ("CSF-6231241A", "Shuho Junmai Daiginjo Black Label", "秀鳳酒造場", "清酒", 2800),
    ("CSF-511l", "Shuho Daiginjo Okuden", "秀鳳酒造場", "清酒", 2500),
    ("CSF-625l", "Shuho Junmai Daiginjo Yamadanishiki 40", "秀鳳酒造場", "清酒", 2300),
    ("CSF-614l", "Shuho Junmai Daiginjo Aiyama", "秀鳳酒造場", "清酒", 2080),
    ("CSF-611l", "Shuho Junmai Daiginjo Dewasansan 33", "秀鳳酒造場", "清酒", 1800),
    ("CSF-612l", "Shuho Junmai Daiginjo Tsuyahime", "秀鳳酒造場", "清酒", 1680),
    ("CSF-618l", "Shuho Junmai Daiginjo Extra DRY", "秀鳳酒造場", "清酒", 1600),
    ("CSF-616l", "Shuho Junmai Daiginjo Yamadanishiki 47", "秀鳳酒造場", "清酒", 1500),
    ("CSF-412l", "Shuho Junmai Ginjo Dewanosato", "秀鳳酒造場", "清酒", 1400),
    ("CSF-411l", "Shuho Junmai Ginjo Beach Side", "秀鳳酒造場", "清酒", 1400),
    ("CSF-207l", "Shuho Junmaishu Spring Sunshine 80", "秀鳳酒造場", "清酒", 1380),
    ("CSF-204l", "Shuho Tokubetsu Junmai Tsuyahime", "秀鳳酒造場", "清酒", 1300),
    ("CSF-242l", "Shuho Tokubetsu Junmai Omachi", "秀鳳酒造場", "清酒", 1300),
    (
        "CSF-245l",
        "Shuho Tokubetsu Junmai Yamahai Dewa Kirari",
        "秀鳳酒造場",
        "清酒",
        1200,
    ),
    ("CSF-141l", "Shuho Honjozo Gainen", "秀鳳酒造場", "清酒", 1200),
    ("CAB-622I", "AMABUKI Junmai Daiginjo Aiyama", "天吹酒造", "清酒", 2980),
    ("CAB-611l", "Amabuki Junmai Daiginjo Kimoto", "天吹酒造", "清酒", 2000),
    ("CAB-623I", "AMABUKI Junmai Daiginjo Banana Kobonama", "天吹酒造", "清酒", 1900),
    ("CAB-614l", "Amabuki Junmai Daiginjo Ringo", "天吹酒造", "清酒", 1800),
    ("CAB-412l", "Amabuki Junmai Ginjo Ichigo", "天吹酒造", "清酒", 1800),
    ("CAB-413I", "AMABUKI Junmai Ginjo Ichigo (bottle)", "天吹酒造", "清酒", 1800),
    ("CAB-620I", "Amabuki Junmai Daiginjo Denim", "天吹酒造", "清酒", 1680),
    ("CAB-621l", "AMABUKI Junmai Daiginjo Linen", "天吹酒造", "清酒", 1680),
    ("CAB-912I", "Amabuki Junmai Cloud Cloudy", "天吹酒造", "清酒", 1680),
    ("CAB-616I", "Amabuki Junmai Daiginjo Koisuru Haru No", "天吹酒造", "清酒", 1680),
    (
        "CAB-618I",
        "Amabuki Junmai Daiginjo Huyunikoisuru Nama",
        "天吹酒造",
        "清酒",
        1600,
    ),
    ("CAB-418I", "Amabuki Junmai Ginjo Knit", "天吹酒造", "清酒", 1580),
    ("CAB-242I", "Amabuki Tokubetsu Junmai Fresh Green", "天吹酒造", "清酒", 1600),
    ("CAB-202I", "Amabuki Ginnokurena", "天吹酒造", "清酒", 1500),
    ("CAB-613I", "Amabuki Junmai Daiginjo 50", "天吹酒造", "清酒", 1500),
    ("CAB-241I", "Amabuki Tokubetsu Junmai Natsunikoisuru", "天吹酒造", "清酒", 1600),
    ("CAB-203I", "Amabuki Junmai Akinikoisuru", "天吹酒造", "清酒", 1400),
    ("CKN-611I", "Kinryo Junmai Daiginjo Kirameki", "西野金陵", "清酒", 4800),
    ("CKN-512I", "Kinryo Daiginjo GOLD Junkinpakuiri", "西野金陵", "清酒", 2980),
    ("CKN-612I", "Kinryo Junmai Daiginjo Yamadanishiki", "西野金陵", "清酒", 2650),
    ("CKN-412I", "Kinryo Junmai Ginjo Setouchi Olive", "西野金陵", "清酒", 2000),
    ("CKN-413I", "Kinryo Junmai Ginjo Koiai", "西野金陵", "清酒", 1550),
    ("CKN-311I", "Kinryo Ginjo Geppaku", "西野金陵", "清酒", 1350),
    ("CKN-242I", "Kinryo Tokubetsu Junmai Chitosemidori", "西野金陵", "清酒", 1150),
    ("CKN-142I", "Kinryo Honjozo Shinku", "西野金陵", "清酒", 980),
    ("CKN-711I", "Kinryo Chokarakuchi", "西野金陵", "清酒", 950),
    ("CYN-411M red", "Marumasu Yonetsuru Junmai Ginjo RED", "米鶴酒造", "清酒", 2450),
    ("CYN-411M grn", "Marumasu Yonetsuru Junmai Ginjo GREEN", "米鶴酒造", "清酒", 2450),
    ("CYN-201I", "CANTABILE", "米鶴酒造", "清酒", 1300),
    # ── 燒酎 ──
    ("BAB-301I", "Amabuki Ginjo Kasutori Shochu", "天吹酒造", "燒酎", 1000),
    ("BDE-201M", "Den En Imo Kuro", "田苑酒造", "燒酎", 1600),
    ("BDE-101F", "Den En Paku Chii Spirit", "田苑酒造", "燒酎", 1580),
    ("BDE-202H", "Den En Envelhecida", "田苑酒造", "燒酎", 1400),
    ("BDE-203T", "Den En Flavor Apple", "田苑酒造", "燒酎", 1280),
    ("BDE-102T", "Den En Flavor Banana", "田苑酒造", "燒酎", 1280),
    ("BDE-103I", "Den En Kin Long Aged", "田苑酒造", "燒酎", 900),
    # ── 利口酒 ──
    ("DAB-U13F", "Amabuki Ryuoh Daiginjo Umeshu", "天吹酒造", "利口酒", 5200),
    ("DAB-U12I", "Amabuki TETE Japan Botanicals", "天吹酒造", "利口酒", 1350),
    ("DAB-U15I", "Amabuki Blood Orange Apollon", "天吹酒造", "利口酒", 1180),
    ("DKN-U11I", "Kinryo Shiroshitato Umeshu", "西野金陵", "利口酒", 1480),
    ("DYG-U01I", "Umeshu Hu Liqueur", "北岡本店", "利口酒", 1150),
    ("DYG-O11I", "Yoshino Monogatari Unshu Mikan", "北岡本店", "利口酒", 1150),
    ("DYG-R11I", "Kitanoyoichi Ringo", "北岡本店", "利口酒", 1150),
    ("DYG-B12I", "Yoshino Monogatari Berry Berry", "北岡本店", "利口酒", 1150),
    # ── 其他 ──
    ("AGJ-691FNV1A", "Costa Russi Grappa NV", "GAJA", "其他", 3050),
    ("AGJ-692FNV1A", "Gaja & Rey Grappa NV", "GAJA", "其他", 3050),
    ("AGJ-693FNV1A", "Sperss Grappa NV", "GAJA", "其他", 3050),
    ("AFC-691HNV1A", "CastelGiocondo Grappa", "TENUTA CASTELGIOCONDO", "其他", 3450),
    ("AVV-364F141A", "TOKAJI ASZÚ, 5 PUTTONYOS '14", "TOKAJ-OREMUS", "其他", 4980),
    ("AVV-362F131A", "TOKAJI ASZÚ, 6 PUTTONYOS '13", "TOKAJ-OREMUS", "其他", 6000),
    # ── 日本白酒 ──
    ("AAY-521J", "ASAYA KOUSHU SUR LIE", "Asaya Winery", "日本白酒", 1400),
]

# ── 建立 Workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ============================================================
# SHEET 1：商品目錄
# ============================================================
ws_cat = wb.active
ws_cat.title = "📋 商品目錄"
ws_cat.sheet_view.showGridLines = False
ws_cat.freeze_panes = "A3"

# 標題列 (row 1)
ws_cat.merge_cells("A1:H1")
title_cell = ws_cat["A1"]
title_cell.value = "⭐ SERGIO VALENTE 星坊酒業 — 商品目錄"
title_cell.fill = fill(C_HEADER)
title_cell.font = Font(name="微軟正黑體", size=14, bold=True, color=C_LIGHTGOLD)
title_cell.alignment = center()
ws_cat.row_dimensions[1].height = 36

# 欄位標題 (row 2)
headers = [
    "商品代碼",
    "商品名稱",
    "品牌/酒廠",
    "類別",
    "售價(NT$)",
    "庫存量",
    "備註",
    "上架狀態",
]
col_widths = [18, 52, 30, 10, 12, 10, 20, 10]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws_cat.cell(row=2, column=i, value=h)
    c.fill = fill(C_GOLD)
    c.font = Font(name="微軟正黑體", size=10, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()
    ws_cat.column_dimensions[get_column_letter(i)].width = w
ws_cat.row_dimensions[2].height = 24

# 類別色彩對應
CAT_COLORS = {
    "紅酒": ("FFF0F0", "D44"),
    "白酒": ("F0F8FF", "48A"),
    "粉紅酒": ("FFF0F8", "D68"),
    "氣泡/香檳": ("FFFBE6", "AA6"),
    "清酒": ("F0FFF4", "3A6"),
    "燒酎": ("F5F5F5", "888"),
    "利口酒": ("F8F0FF", "86A"),
    "其他": ("FAFAFA", "999"),
    "日本白酒": ("F0FFFF", "49A"),
}

# 填入商品資料
for row_i, (code, name, brand, cat, price) in enumerate(PRODUCTS, 3):
    bg, _ = CAT_COLORS.get(cat, ("FFFFFF", "666"))
    row_data = [code, name, brand, cat, price, "", "", "上架中"]
    for col_i, val in enumerate(row_data, 1):
        c = ws_cat.cell(row=row_i, column=col_i, value=val)
        c.fill = fill(bg if col_i > 1 else C_WHITE)
        c.font = font(size=9)
        c.border = border_thin()
        if col_i == 1:
            c.font = Font(name="Consolas", size=8, color="666666")
            c.alignment = center()
        elif col_i == 2:
            c.alignment = left_middle()
        elif col_i in (3, 4, 7, 8):
            c.alignment = center()
        elif col_i == 5:
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0"
        elif col_i == 6:
            c.alignment = center()
            c.number_format = "0"
        elif col_i == 8:
            c.alignment = center()
            c.font = Font(
                name="微軟正黑體",
                size=8,
                bold=True,
                color=C_GREEN if val == "上架中" else C_RED,
            )
    ws_cat.row_dimensions[row_i].height = 18

# 下拉驗證：上架狀態
dv_status = DataValidation(
    type="list", formula1='"上架中,暫停供應,已停產"', allow_blank=False
)
dv_status.sqref = f"H3:H{2+len(PRODUCTS)}"
ws_cat.add_data_validation(dv_status)

# 下拉驗證：類別
cat_list = ",".join(sorted(set(p[3] for p in PRODUCTS)))
dv_cat = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=True)
dv_cat.sqref = f"D3:D{2+len(PRODUCTS)}"
ws_cat.add_data_validation(dv_cat)

# 條件格式：售價顏色刻度
last_row = 2 + len(PRODUCTS)
ws_cat.conditional_formatting.add(
    f"E3:E{last_row}",
    ColorScaleRule(
        start_type="min",
        start_color="63BE7B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="F8696B",
    ),
)

# ============================================================
# SHEET 2：出貨單
# ============================================================
ws_ship = wb.create_sheet("📦 出貨單")
ws_ship.sheet_view.showGridLines = False
ws_ship.freeze_panes = "A5"

# 大標題
ws_ship.merge_cells("A1:L1")
h = ws_ship["A1"]
h.value = "📦 SERGIO VALENTE 星坊酒業 — 出貨單"
h.fill = fill(C_HEADER)
h.font = Font(name="微軟正黑體", size=13, bold=True, color=C_LIGHTGOLD)
h.alignment = center()
ws_ship.row_dimensions[1].height = 34

# 出貨資訊列
info = [
    ("A2", "出貨日期"),
    ("B2", str(date.today())),
    ("D2", "出貨單號"),
    ("E2", "SV-2026-001"),
    ("G2", "客戶名稱"),
    ("H2", ""),
    ("J2", "業務人員"),
    ("K2", ""),
]
for cell_ref, val in info:
    c = ws_ship[cell_ref]
    c.value = val
    c.font = Font(
        name="微軟正黑體",
        size=9,
        bold=("日期" in val or "單號" in val or "名稱" in val or "人員" in val),
    )
    c.alignment = left_middle()
ws_ship.row_dimensions[2].height = 20

ws_ship.merge_cells("A3:L3")
ws_ship["A3"].value = (
    "── 收件資訊 ───────────────────────────────────────────────────────────────────────"
)
ws_ship["A3"].font = Font(name="微軟正黑體", size=8, color="AAAAAA")

# 欄標題 (row 4)
ship_headers = [
    "#",
    "商品代碼",
    "商品名稱",
    "品牌/酒廠",
    "類別",
    "售價(NT$)",
    "數量",
    "單位",
    "小計(NT$)",
    "折扣%",
    "折後小計",
    "備註",
]
ship_widths = [4, 18, 50, 28, 10, 12, 8, 6, 14, 8, 14, 16]
for i, (h_txt, w) in enumerate(zip(ship_headers, ship_widths), 1):
    c = ws_ship.cell(row=4, column=i, value=h_txt)
    c.fill = fill(C_DARKGOLD)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()
    ws_ship.column_dimensions[get_column_letter(i)].width = w
ws_ship.row_dimensions[4].height = 22

# 出貨明細（30列空白可填寫，含公式）
for r in range(5, 36):
    ws_ship.cell(row=r, column=1, value=r - 4).alignment = center()
    ws_ship.cell(row=r, column=1).font = font(size=8, color="AAAAAA")
    for col in range(1, 13):
        c = ws_ship.cell(row=r, column=col)
        c.fill = fill(C_CREAM if r % 2 == 0 else C_WHITE)
        c.border = border_thin()
        c.alignment = center() if col in (1, 5, 6, 7, 8, 9, 10, 11) else left_middle()
    # 小計公式 = 售價 × 數量
    c_sub = ws_ship.cell(row=r, column=9)
    c_sub.value = f'=IF(F{r}="","",F{r}*G{r})'
    c_sub.number_format = "#,##0"
    # 折後小計 = 小計 × (1 - 折扣/100)
    c_disc = ws_ship.cell(row=r, column=11)
    c_disc.value = f'=IF(I{r}="","",IF(J{r}="",I{r},I{r}*(1-J{r}/100)))'
    c_disc.number_format = "#,##0"
    ws_ship.row_dimensions[r].height = 18

# 合計列
total_row = 36
ws_ship.merge_cells(f"A{total_row}:H{total_row}")
ws_ship[f"A{total_row}"].value = "合計"
ws_ship[f"A{total_row}"].font = Font(name="微軟正黑體", size=10, bold=True)
ws_ship[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
ws_ship[f"A{total_row}"].fill = fill(C_LIGHTGOLD)

for col_idx, formula_col in [(9, "I"), (11, "K")]:
    c = ws_ship.cell(row=total_row, column=col_idx)
    c.value = f"=SUM({formula_col}5:{formula_col}35)"
    c.number_format = "#,##0"
    c.font = Font(name="微軟正黑體", size=11, bold=True, color=C_DARKGOLD)
    c.fill = fill(C_LIGHTGOLD)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border_medium()
ws_ship.row_dimensions[total_row].height = 28

# 備註欄
ws_ship.merge_cells(f"A{total_row+2}:L{total_row+3}")
ws_ship[f"A{total_row+2}"].value = "備註："
ws_ship[f"A{total_row+2}"].fill = fill(C_GRAY)
ws_ship[f"A{total_row+2}"].font = Font(name="微軟正黑體", size=9)
ws_ship[f"A{total_row+2}"].alignment = left_middle()

# VLOOKUP 查詢提示（隱藏輔助區）
ws_ship["N1"].value = "▶ 輸入商品代碼可自動帶入"
ws_ship["N1"].font = Font(name="微軟正黑體", size=8, color="AAAAAA", italic=True)

# ============================================================
# SHEET 3：銷售報表
# ============================================================
ws_sales = wb.create_sheet("📊 銷售報表")
ws_sales.sheet_view.showGridLines = False
ws_sales.freeze_panes = "A4"

# 標題
ws_sales.merge_cells("A1:N1")
h = ws_sales["A1"]
h.value = "📊 SERGIO VALENTE 星坊酒業 — 銷售報表"
h.fill = fill(C_HEADER)
h.font = Font(name="微軟正黑體", size=13, bold=True, color=C_LIGHTGOLD)
h.alignment = center()
ws_sales.row_dimensions[1].height = 34

# 篩選條件列
ws_sales["A2"].value = "月份："
ws_sales["B2"].value = str(date.today())[:7]  # YYYY-MM
ws_sales["D2"].value = "業務："
ws_sales["E2"].value = ""
ws_sales["G2"].value = "客戶："
ws_sales["H2"].value = ""
ws_sales["J2"].value = "類別："
ws_sales["K2"].value = "全部"
for cell_ref in ["A2", "D2", "G2", "J2"]:
    ws_sales[cell_ref].font = Font(name="微軟正黑體", size=9, bold=True)
ws_sales.row_dimensions[2].height = 20

# 欄標題
sales_headers = [
    "出貨日期",
    "出貨單號",
    "客戶名稱",
    "商品代碼",
    "商品名稱",
    "品牌/酒廠",
    "類別",
    "單價",
    "數量",
    "小計",
    "折扣%",
    "折後小計",
    "業務",
    "備註",
]
sales_widths = [12, 14, 18, 18, 42, 26, 10, 12, 8, 14, 8, 14, 12, 16]
for i, (h_txt, w) in enumerate(zip(sales_headers, sales_widths), 1):
    c = ws_sales.cell(row=3, column=i, value=h_txt)
    c.fill = fill(C_GOLD)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()
    ws_sales.column_dimensions[get_column_letter(i)].width = w
ws_sales.row_dimensions[3].height = 22

# 500 列空白銷售記錄
for r in range(4, 504):
    alt_bg = C_CREAM if r % 2 == 0 else C_WHITE
    for col in range(1, 15):
        c = ws_sales.cell(row=r, column=col)
        c.fill = fill(alt_bg)
        c.border = border_thin()
        c.font = font(size=9)
        c.alignment = (
            center() if col in (1, 2, 7, 8, 9, 10, 11, 12, 13) else left_middle()
        )
    # 小計公式
    ws_sales.cell(row=r, column=10).value = f'=IF(H{r}="","",H{r}*I{r})'
    ws_sales.cell(row=r, column=10).number_format = "#,##0"
    ws_sales.cell(row=r, column=12).value = (
        f'=IF(J{r}="","",IF(K{r}="",J{r},J{r}*(1-K{r}/100)))'
    )
    ws_sales.cell(row=r, column=12).number_format = "#,##0"
    ws_sales.cell(row=r, column=8).number_format = "#,##0"
    ws_sales.row_dimensions[r].height = 17

# 合計列
sum_row = 504
ws_sales.merge_cells(f"A{sum_row}:G{sum_row}")
ws_sales[f"A{sum_row}"].value = "▶ 期間合計"
ws_sales[f"A{sum_row}"].font = Font(name="微軟正黑體", size=10, bold=True)
ws_sales[f"A{sum_row}"].alignment = Alignment(horizontal="right")
ws_sales[f"A{sum_row}"].fill = fill(C_LIGHTGOLD)
for col_idx, formula_col in [(10, "J"), (12, "L")]:
    c = ws_sales.cell(row=sum_row, column=col_idx)
    c.value = f'=SUMIF(G4:G503,"<>",{formula_col}4:{formula_col}503)'
    c.number_format = "#,##0"
    c.font = Font(name="微軟正黑體", size=11, bold=True, color=C_DARKGOLD)
    c.fill = fill(C_LIGHTGOLD)
    c.alignment = Alignment(horizontal="right")
    c.border = border_medium()
ws_sales.row_dimensions[sum_row].height = 26

# ============================================================
# SHEET 4：統計儀表板
# ============================================================
ws_dash = wb.create_sheet("📈 統計儀表板")
ws_dash.sheet_view.showGridLines = False

# 標題
ws_dash.merge_cells("A1:P1")
h = ws_dash["A1"]
h.value = "📈 銷售統計儀表板"
h.fill = fill(C_HEADER)
h.font = Font(name="微軟正黑體", size=14, bold=True, color=C_LIGHTGOLD)
h.alignment = center()
ws_dash.row_dimensions[1].height = 36

# ── 類別統計小表 ──
ws_dash["A3"].value = "類別銷售統計"
ws_dash["A3"].font = Font(name="微軟正黑體", size=11, bold=True, color=C_DARKGOLD)
ws_dash["A3"].fill = fill(C_LIGHTGOLD)

stat_headers = ["類別", "商品數", "合計銷售額(NT$)", "平均售價(NT$)"]
for i, h_txt in enumerate(stat_headers, 1):
    c = ws_dash.cell(row=4, column=i, value=h_txt)
    c.fill = fill(C_GOLD)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color=C_WHITE)
    c.alignment = center()
    c.border = border_thin()
    ws_dash.column_dimensions[get_column_letter(i)].width = [14, 10, 22, 18][i - 1]

categories = [
    "紅酒",
    "白酒",
    "粉紅酒",
    "氣泡/香檳",
    "清酒",
    "燒酎",
    "利口酒",
    "其他",
    "日本白酒",
]
cat_counts = {cat: sum(1 for p in PRODUCTS if p[3] == cat) for cat in categories}
cat_prices = {cat: [p[4] for p in PRODUCTS if p[3] == cat] for cat in categories}

for row_i, cat in enumerate(categories, 5):
    count = cat_counts[cat]
    prices_list = cat_prices[cat]
    avg_price = sum(prices_list) / len(prices_list) if prices_list else 0
    bg, _ = CAT_COLORS.get(cat, ("FFFFFF", "666"))
    row_data = [
        cat,
        count,
        f"=SUMIF('📊 銷售報表'!G$4:G$503,\"{cat}\",'📊 銷售報表'!L$4:L$503)",
        round(avg_price),
    ]
    for col_i, val in enumerate(row_data, 1):
        c = ws_dash.cell(row=row_i, column=col_i, value=val)
        c.fill = fill(bg)
        c.border = border_thin()
        c.font = font(size=9)
        c.alignment = center()
        if col_i == 3:
            c.number_format = "#,##0"
        if col_i == 4:
            c.number_format = "#,##0"
    ws_dash.row_dimensions[row_i].height = 18

# 總計列
total_r = 5 + len(categories)
ws_dash.cell(row=total_r, column=1, value="合計").font = Font(
    name="微軟正黑體", size=9, bold=True
)
ws_dash.cell(row=total_r, column=2, value=len(PRODUCTS)).font = Font(
    name="微軟正黑體", size=9, bold=True
)
c_total = ws_dash.cell(row=total_r, column=3, value=f"=SUM(C5:C{total_r-1})")
c_total.number_format = "#,##0"
c_total.font = Font(name="微軟正黑體", size=10, bold=True, color=C_DARKGOLD)
for col in range(1, 5):
    ws_dash.cell(row=total_r, column=col).fill = fill(C_LIGHTGOLD)
    ws_dash.cell(row=total_r, column=col).border = border_medium()
ws_dash.row_dimensions[total_r].height = 22

# ── 月度業績 KPI ──
ws_dash["F3"].value = "月度業績 KPI"
ws_dash["F3"].font = Font(name="微軟正黑體", size=11, bold=True, color=C_DARKGOLD)
ws_dash["F3"].fill = fill(C_LIGHTGOLD)
ws_dash.merge_cells("F3:L3")

kpis = [
    (
        "本月銷售總額",
        f"=SUMIF('📊 銷售報表'!A$4:A$503,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'📊 銷售報表'!L$4:L$503)",
        C_GREEN,
    ),
    (
        "本月出貨筆數",
        f"=COUNTIF('📊 銷售報表'!A$4:A$503,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1))",
        C_GOLD,
    ),
    ("最高單筆金額", "=MAX('📊 銷售報表'!L$4:L$503)", C_DARKGOLD),
    ("平均客單價", f"=IFERROR(AVERAGE('📊 銷售報表'!L$4:L$503),0)", C_BLACK),
]
for ki, (label, formula, color) in enumerate(kpis):
    r = 5 + ki * 3
    ws_dash.merge_cells(f"F{r}:L{r}")
    ws_dash.merge_cells(f"F{r+1}:L{r+1}")
    lc = ws_dash.cell(row=r, column=6, value=label)
    lc.font = Font(name="微軟正黑體", size=8, color="777777")
    lc.alignment = center()
    lc.fill = fill(C_GRAY)
    vc = ws_dash.cell(row=r + 1, column=6, value=formula)
    vc.font = Font(name="微軟正黑體", size=18, bold=True, color=color)
    vc.alignment = center()
    vc.fill = fill(C_CREAM)
    vc.number_format = "#,##0"
    for col in range(6, 13):
        ws_dash.cell(row=r, column=col).border = border_thin()
        ws_dash.cell(row=r + 1, column=col).border = border_thin()
    ws_dash.row_dimensions[r].height = 16
    ws_dash.row_dimensions[r + 1].height = 30
    ws_dash.column_dimensions[get_column_letter(6)].width = 8

for c in range(6, 13):
    ws_dash.column_dimensions[get_column_letter(c)].width = 7

# ── 商品查詢工具提示 ──
ws_dash["A22"].value = "💡 使用說明"
ws_dash["A22"].font = Font(name="微軟正黑體", size=10, bold=True, color=C_DARKGOLD)
ws_dash.merge_cells("A22:P22")
ws_dash["A22"].fill = fill(C_LIGHTGOLD)

tips = [
    "1. 【商品目錄】：星坊酒業全品項 + 庫存管理，可篩選類別/狀態",
    "2. 【出貨單】：每次出貨填寫，輸入售價與數量，小計/合計自動計算",
    "3. 【銷售報表】：按日累積銷售紀錄，共 500 列，小計與折後金額自動帶入",
    "4. 【統計儀表板】：自動從銷售報表彙整各類別銷售額 & 月度 KPI",
    "5. 篩選訣竅：在商品目錄選「資料 → 篩選」，可按類別/狀態篩選商品",
]
for ti, tip in enumerate(tips):
    r = 23 + ti
    ws_dash.merge_cells(f"A{r}:P{r}")
    c = ws_dash.cell(row=r, column=1, value=tip)
    c.font = Font(name="微軟正黑體", size=9)
    c.alignment = left_middle()
    c.fill = fill(C_CREAM if ti % 2 == 0 else C_WHITE)
    ws_dash.row_dimensions[r].height = 17

# ── 標籤欄寬 ──
for col in "ABCDEFGHIJKLMNOP":
    if ws_dash.column_dimensions[col].width < 6:
        ws_dash.column_dimensions[col].width = 6

# ============================================================
# 全域設定：列印紙張
# ============================================================
for ws in [ws_cat, ws_ship, ws_sales]:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

# ── 設定 Tab 顏色 ──
ws_cat.sheet_properties.tabColor = "B7874A"
ws_ship.sheet_properties.tabColor = "3D7A52"
ws_sales.sheet_properties.tabColor = "2E75B6"
ws_dash.sheet_properties.tabColor = "8B4513"

# ── 儲存 ──
output_path = "/Users/lien/Downloads/星坊酒業_販售清單.xlsx"
wb.save(output_path)
print(f"✅ 已儲存：{output_path}")
print(f"   商品目錄：{len(PRODUCTS)} 筆")
print(f"   出貨單：30 列可填入")
print(f"   銷售報表：500 列可填入")
print(f"   統計儀表板：自動彙整")
