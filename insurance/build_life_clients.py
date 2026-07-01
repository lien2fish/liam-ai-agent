#!/usr/bin/env python3
"""從連傳正客戶清單抓出保單類型含「壽」的客戶，存本機 xlsx+numbers，並建 Notion 名單。

本機檔保留完整欄位（含身分證）；Notion 不含身分證。
"""
import json, os, urllib.request
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = "/Users/lien/Desktop/20260701_連傳正_客戶清單.xlsx"
OUTDIR = "/Users/lien/Desktop/鉅鑫管理顧問/磊山保經/客戶名單"
STAMP = "20260701"
COLS = [
    "保險年齡狀態",
    "姓名",
    "身分證號碼",
    "生日",
    "年齡",
    "保險年齡增加日",
    "保險年齡",
    "保單類型",
    "手機/電話",
    "戶籍地址",
]


def extract():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["20260701_連傳正_客戶清單"]
    hdr = [c.value for c in ws[1]]
    ti = hdr.index("保單類型")
    rows = []
    for r in range(2, ws.max_row + 1):
        ptype = ws.cell(r, ti + 1).value
        if ptype and "壽" in str(ptype):
            rows.append([ws.cell(r, i + 1).value for i in range(len(COLS))])
    return rows


def to_isodate(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10] if v else ""


def save_xlsx(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "壽險客戶名單"
    navy = PatternFill("solid", fgColor="1F4E79")
    for c, h in enumerate(COLS, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = navy
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    ws.freeze_panes = "A2"
    widths = [14, 12, 14, 12, 6, 16, 10, 12, 14, 34]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    wb.save(path)


def save_numbers(rows, path):
    from numbers_parser import Document

    doc = Document()
    table = doc.sheets[0].tables[0]
    need_r, need_c = len(rows) + 1, len(COLS)
    while table.num_rows < need_r:
        table.add_row()
    while table.num_cols < need_c:
        table.add_column()
    for c, h in enumerate(COLS):
        table.write(0, c, h)
    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row):
            table.write(
                r,
                c,
                "" if v is None else (v if isinstance(v, (int, float)) else str(v)),
            )
    doc.save(path)


def build_notion(rows):
    tok = open(os.path.expanduser("~/.config/notion_token")).read().strip()
    H = {
        "Authorization": f"Bearer {tok}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json",
    }
    CRM = "358f4149-a6aa-8088-9e6d-f5361d05cd12"

    def api(method, path, body):
        req = urllib.request.Request(
            f"https://api.notion.com/v1{path}",
            data=json.dumps(body).encode(),
            headers=H,
            method=method,
        )
        return json.load(urllib.request.urlopen(req))

    props = {
        "姓名": {"title": {}},
        "保單類型": {
            "select": {
                "options": [
                    {"name": "壽", "color": "blue"},
                    {"name": "壽產", "color": "green"},
                    {"name": "壽團", "color": "purple"},
                    {"name": "壽產團", "color": "orange"},
                ]
            }
        },
        "手機電話": {"phone_number": {}},
        "年齡": {"number": {"format": "number"}},
        "生日": {"date": {}},
        "保險年齡": {"number": {"format": "number"}},
        "保險年齡增加日": {"date": {}},
        "保險年齡狀態": {
            "select": {"options": [{"name": "即將+1歲", "color": "yellow"}]}
        },
        "戶籍地址": {"rich_text": {}},
        "拜訪週期": {
            "select": {
                "options": [
                    {"name": "每季", "color": "red"},
                    {"name": "每半年", "color": "green"},
                    {"name": "每年", "color": "gray"},
                ]
            }
        },
        "上次拜訪日": {"date": {}},
        "下次拜訪日": {"date": {}},
        "拜訪狀態": {
            "select": {
                "options": [
                    {"name": "待拜訪", "color": "orange"},
                    {"name": "已完成", "color": "green"},
                    {"name": "暫緩", "color": "gray"},
                ]
            }
        },
        "拜訪備註": {"rich_text": {}},
    }
    db = api(
        "POST",
        "/databases",
        {
            "parent": {"type": "page_id", "page_id": CRM},
            "title": [
                {"type": "text", "text": {"content": "🛡️ 壽險客戶名單（固定拜訪）"}}
            ],
            "properties": props,
        },
    )
    dbid = db["id"]
    print(f"  Notion DB 建立：{dbid}")

    def rt(s):
        return [{"type": "text", "text": {"content": str(s)[:1900]}}] if s else []

    for row in rows:
        st, name, _id, bday, age, incd, iage, ptype, phone, addr = row
        p = {
            "姓名": {"title": rt(name or "（未命名）")},
            "保單類型": {"select": {"name": str(ptype)}},
            "手機電話": {"phone_number": str(phone) if phone else None},
            "戶籍地址": {"rich_text": rt(addr)},
            "拜訪週期": {"select": {"name": "每半年"}},
            "拜訪狀態": {"select": {"name": "待拜訪"}},
        }
        if isinstance(age, (int, float)):
            p["年齡"] = {"number": int(age)}
        if isinstance(iage, (int, float)):
            p["保險年齡"] = {"number": int(iage)}
        if to_isodate(bday):
            p["生日"] = {"date": {"start": to_isodate(bday)}}
        if to_isodate(incd):
            p["保險年齡增加日"] = {"date": {"start": to_isodate(incd)}}
        if st and str(st).strip():
            p["保險年齡狀態"] = {"select": {"name": str(st).strip()}}
        api("POST", "/pages", {"parent": {"database_id": dbid}, "properties": p})
    return dbid


def main():
    rows = extract()
    print(f"擷取壽險客戶 {len(rows)} 位")
    os.makedirs(OUTDIR, exist_ok=True)
    xp = f"{OUTDIR}/壽險客戶名單_{STAMP}.xlsx"
    save_xlsx(rows, xp)
    print(f"✅ xlsx：{xp}")
    try:
        npath = f"{OUTDIR}/壽險客戶名單_{STAMP}.numbers"
        save_numbers(rows, npath)
        print(f"✅ numbers：{npath}")
    except Exception as e:
        print(f"⚠️ numbers 生成失敗（稍後改用 AppleScript 轉）：{e}")
    dbid = build_notion(rows)
    json.dump(
        {"life_clients_db": dbid, "reminder_days": 7},
        open(
            os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "visit_config.json"
            ),
            "w",
        ),
        ensure_ascii=False,
        indent=2,
    )
    print(f"✅ Notion 匯入完成，DB id 已存 visit_config.json")


if __name__ == "__main__":
    main()
