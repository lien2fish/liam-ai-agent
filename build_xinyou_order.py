#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
建立鑫酒藏整合出貨單
- 出貨單：客戶下拉 + 品項下拉 + VLOOKUP自動帶入
- 商品目錄：完整品項清單（可新增）
- 客戶名單：完整客戶清單（可新增）
- 銷售紀錄：歷史訂單記錄
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME

OUTPUT = "/Users/lien/Desktop/鉅鑫管理顧問/鑫酒藏/鑫酒藏品項＆客戶紀錄單.xlsx"

# ── 商品目錄 ────────────────────────────────────────────────
PRODUCTS = [
    ("WGW010201", "教父酒莊鑽石低卡灰皮諾白酒", "白酒", 1100, "瓶"),
    ("WGW010202", "教父酒莊大導演白酒", "白酒", 1950, "瓶"),
    ("WGR010203", "教父酒莊大導演卡本內紅酒", "紅酒", 2250, "瓶"),
    ("WGR010204", "教父酒莊大導演黑皮諾紅酒", "紅酒", 2150, "瓶"),
    ("WGR010205", "教父酒莊黑鑽", "紅酒", 1550, "瓶"),
    ("WGR010301", "伯恩一級園命運獨佔園紅酒", "紅酒", 4000, "瓶"),
    ("WGR010302", "梅克雷紅酒", "紅酒", 1500, "瓶"),
    ("WGR010303", "佩南維哲雷一級園紅酒", "紅酒", None, "瓶"),
    ("WGW010304", "伯恩一級園蜂蜜園白酒", "白酒", None, "瓶"),
    ("WGR010401", "Caymus50周年禮盒", "紅酒", 3250, "瓶"),
    ("WGR010501", "CONTIUUM心傳紅葡萄酒2022", "紅酒", 13580, "瓶"),
    ("WGR010502", "NOVICIUM初心紅葡萄酒2022", "紅酒", 8120, "瓶"),
    ("WGW010503", "SENTIUM從心蘇維儂白酒2023", "白酒", 3220, "瓶"),
    ("WGR030101", "MASI蛇年紀念酒", "紅酒", None, "瓶"),
    ("WGR030102", "MASI馬年紀念酒", "紅酒", None, "瓶"),
    ("WGR030103", "MASI花之園紅酒", "紅酒", 1450, "瓶"),
    ("WGW030104", "瑪西安可白酒", "白酒", 1200, "瓶"),
    ("WGR030105", "MASI Amarone", "紅酒", 3800, "瓶"),
    ("WGS030201", "卡內維頂級不甜氣泡酒", "氣泡酒", 1500, "瓶"),
    ("W110101", "皇家禮炮", "威士忌", 2600, "瓶"),
    ("W110701", "大摩15年", "威士忌", 2600, "瓶"),
    ("W110702", "大摩18年", "威士忌", 6700, "瓶"),
    ("W110703", "大摩1L", "威士忌", 2000, "瓶"),
    ("WGS020101", "RARE 2008", "香檳", 12000, "瓶"),
    ("WGS020201", "沛芙希夢一號黑中白", "香檳", 7000, "瓶"),
    ("WGS020301", "Bollinger NV", "香檳", 2000, "瓶"),
    ("W1401", "清酒", "清酒", 2000, "瓶"),
    ("S026", "梅酒", "梅酒", None, "瓶"),
]

# ── 客戶名單 ────────────────────────────────────────────────
CUSTOMERS = [
    ("徐培鈞", "0935-290078", "桃園市大園區中正東路一段840巷97號"),
    ("kevin", "", "台北市中山區長春路368號4樓"),
    ("黃明偉", "0932-112379", "桃園市大園區大業路二段236號8樓"),
    ("Stela", "", ""),
    ("Paker", "", ""),
    ("闕育美", "", "台北市內湖區新明路253號"),
    ("許立誼", "", ""),
    ("周明智", "", "台北市中山區吉林路6號14樓之1"),
    ("徐睿煜Chris", "", "新北市新店區安祥路156巷6號1樓"),
    ("楊曛聰", "0912-598179", "新北市板橋區文化路一段362號34樓之2"),
    ("何祖舜", "", "台北市中正區衡陽路7號13樓"),
    ("張萬成", "0936-066048", "新北市林口區文化三路一段462號4樓"),
    ("李青齡", "0918-822805", "新北市永和區仁愛路250號26F"),
    ("鄭雅玲", "0935-058588", "台北市大安區建國南路一段212巷51號8樓"),
    ("朱日銓", "", ""),
    ("許兆慶", "", ""),
    ("張家齊", "", "新北市蘆洲區長榮路339號5樓"),
    ("郭鑒之", "", "台北市中山區中山三段26號5樓之1"),
    ("蕭有為", "", "台北市中山區中山三段26號5樓之1"),
    ("彭鼎堯", "", "新北市中和區圓通路369巷30弄46號"),
    ("施汎泉", "", "台北市大安區羅斯福路二段105號6樓之1"),
    ("林美瑄", "2995-3365", "新北市三重區光復路一段81號10樓"),
    ("林志憶", "", "台北市內湖區行善路417巷28號10樓"),
    ("車普羅", "", ""),
    ("陳怡君", "0922-715334", "台北市中山北路二段65巷10號8樓"),
    ("劉仁閔", "", "台北市仁愛路二段71號4樓之6"),
    ("鄭昱仁", "", "台北市松山區光復南路46巷36號2樓之1"),
    ("王淑慧", "", "台北市民權東路六段2號7樓"),
    ("肪煎商行 鄭興安", "0922-050113", "新北市新莊區福德二街51號"),
]

# ── 樣式 ────────────────────────────────────────────────────
GOLD = "B8860B"
DARK = "1A1A2E"
LIGHT_GOLD = "FFF8E7"
GRAY = "F5F5F5"
MID_GRAY = "DDDDDD"

def font(bold=False, size=11, color="000000", name="微軟正黑體"):
    return Font(bold=bold, size=size, color=color, name=name)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    s = Side(style="thin", color=MID_GRAY)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    b = Border()
    for side in sides:
        setattr(b, side, s)
    return b

def build():
    wb = openpyxl.Workbook()

    # ── 工作表順序 ──────────────────────────────────────────
    ws_catalog = wb.active
    ws_catalog.title = "商品目錄"
    ws_cust    = wb.create_sheet("客戶名單")
    ws_sales   = wb.create_sheet("銷售紀錄")

    build_catalog(ws_catalog)
    build_customers(ws_cust)
    build_sales(ws_sales)

    wb.save(OUTPUT)
    print(f"✓ 已儲存：{OUTPUT}")

# ── 商品目錄 ────────────────────────────────────────────────
def build_catalog(ws):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8

    headers = ["編號", "品項名稱", "類別", "單價", "單位"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = font(bold=True, color="FFFFFF")
        cell.fill = fill(DARK)
        cell.alignment = align("center")
        cell.border = thin_border()

    ws.row_dimensions[1].height = 22

    for r, (code, name, cat, price, unit) in enumerate(PRODUCTS, 2):
        row_fill = fill(LIGHT_GOLD) if r % 2 == 0 else fill("FFFFFF")
        data = [code, name, cat, price, unit]
        for c, v in enumerate(data, 1):
            cell = ws.cell(r, c, v)
            cell.fill = row_fill
            cell.border = thin_border()
            cell.alignment = align("center" if c != 2 else "left")
            if c == 4 and v:
                cell.number_format = "#,##0"

    # 凍結表頭
    ws.freeze_panes = "A2"

# ── 客戶名單 ────────────────────────────────────────────────
def build_customers(ws):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20

    headers = ["姓名", "電話", "地址", "Email", "備註"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = font(bold=True, color="FFFFFF")
        cell.fill = fill(DARK)
        cell.alignment = align("center")
        cell.border = thin_border()

    ws.row_dimensions[1].height = 22

    for r, (name, phone, addr) in enumerate(CUSTOMERS, 2):
        row_fill = fill(LIGHT_GOLD) if r % 2 == 0 else fill("FFFFFF")
        for c, v in enumerate([name, phone, addr, "", ""], 1):
            cell = ws.cell(r, c, v)
            cell.fill = row_fill
            cell.border = thin_border()
            cell.alignment = align("left")

    ws.freeze_panes = "A2"

# ── 出貨單 ────────────────────────────────────────────────────
# 精簡排版：移除空白列，適合 A4 單頁列印
# 列配置：1標題 2分隔 3訂單資訊 4客戶資訊 5地址 6分隔 7品項表頭 8~19品項 20小計 21折扣 22運費 23總計 24備註 25匯款標題 26~28匯款明細 29警語
ITEM_START = 8
ITEM_ROWS  = 12

def build_order(ws):
    from openpyxl.worksheet.page import PageMargins

    # ── 欄寬 ──
    col_widths = {"A": 13, "B": 32, "C": 9, "D": 7, "E": 7, "F": 11, "G": 13}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 1：標題 ──
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "鑫  酒  藏  出  貨  單"
    t.font = Font(bold=True, size=16, color=GOLD, name="微軟正黑體")
    t.fill = fill(DARK)
    t.alignment = align("center", "center")
    ws.row_dimensions[1].height = 36

    # ── Row 2：金色分隔 ──
    ws.merge_cells("A2:G2")
    ws["A2"].fill = fill(GOLD)
    ws.row_dimensions[2].height = 3

    # ── Row 3：訂單編號 + 日期 ──
    _label(ws, "A3", "訂單編號")
    ws.merge_cells("B3:C3")
    _input(ws, "B3", "")
    _label(ws, "E3", "訂單日期")
    ws.merge_cells("F3:G3")
    d = ws["F3"]
    d.value = ""
    d.border = thin_border()
    d.font = font(size=11)
    d.alignment = align("center")
    ws.row_dimensions[3].height = 20

    # ── Row 4：客戶姓名 + 電話 ──
    _label(ws, "A4", "客戶姓名")
    ws.merge_cells("B4:C4")
    cust_cell = ws["B4"]
    cust_cell.value = ""
    cust_cell.border = thin_border()
    cust_cell.font = font(size=11, bold=True, color=DARK)
    cust_cell.alignment = align("left")
    cust_cell.fill = fill(LIGHT_GOLD)

    _label(ws, "E4", "電　　話")
    ws.merge_cells("F4:G4")
    ws["F4"].value = '=IFERROR(VLOOKUP(B4,客戶名單!$A:$C,2,FALSE),"")'
    ws["F4"].border = thin_border()
    ws["F4"].font = font(size=11)
    ws["F4"].alignment = align("left")
    ws.row_dimensions[4].height = 20

    # ── Row 5：地址 ──
    _label(ws, "A5", "地　　址")
    ws.merge_cells("B5:G5")
    ws["B5"].value = '=IFERROR(VLOOKUP(B4,客戶名單!$A:$C,3,FALSE),"")'
    ws["B5"].border = thin_border()
    ws["B5"].font = font(size=11)
    ws["B5"].alignment = align("left")
    ws.row_dimensions[5].height = 20

    # ── Row 6：金色分隔 ──
    ws.merge_cells("A6:G6")
    ws["A6"].fill = fill(GOLD)
    ws.row_dimensions[6].height = 3

    # ── Row 7：品項表頭 ──
    item_headers = ["品項編號", "品　項　名　稱", "類別", "數量", "單位", "單價", "金額"]
    for c, h in enumerate(item_headers, 1):
        cell = ws.cell(7, c, h)
        cell.font = font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill(DARK)
        cell.alignment = align("center")
        cell.border = thin_border()
    ws.row_dimensions[7].height = 20

    # ── Rows 8~19：品項明細 ──
    for i in range(ITEM_ROWS):
        r = ITEM_START + i
        row_fill = fill(LIGHT_GOLD) if i % 2 == 0 else fill("FFFFFF")
        ws.row_dimensions[r].height = 20

        a = ws.cell(r, 1)
        a.value = f'=IFERROR(INDEX(商品目錄!$A:$A,MATCH(B{r},商品目錄!$B:$B,0)),"")'
        a.border = thin_border()
        a.fill = fill(GRAY)
        a.alignment = align("center")
        a.font = font(size=9, color="666666")

        b = ws.cell(r, 2)
        b.value = ""
        b.border = thin_border()
        b.fill = row_fill
        b.alignment = align("left")
        b.font = font(size=10)

        c_cell = ws.cell(r, 3)
        c_cell.value = f'=IFERROR(VLOOKUP(B{r},商品目錄!$B:$C,2,FALSE),"")'
        c_cell.border = thin_border()
        c_cell.fill = fill(GRAY)
        c_cell.alignment = align("center")
        c_cell.font = font(size=9, color="666666")

        d_cell = ws.cell(r, 4)
        d_cell.value = None
        d_cell.border = thin_border()
        d_cell.fill = row_fill
        d_cell.alignment = align("center")
        d_cell.font = font(size=10)

        e_cell = ws.cell(r, 5)
        e_cell.value = f'=IFERROR(VLOOKUP(B{r},商品目錄!$B:$E,4,FALSE),"")'
        e_cell.border = thin_border()
        e_cell.fill = fill(GRAY)
        e_cell.alignment = align("center")
        e_cell.font = font(size=9, color="666666")

        f_cell = ws.cell(r, 6)
        f_cell.value = f'=IFERROR(VLOOKUP(B{r},商品目錄!$B:$D,3,FALSE),"")'
        f_cell.border = thin_border()
        f_cell.fill = row_fill
        f_cell.alignment = align("center")
        f_cell.font = font(size=10)
        f_cell.number_format = "#,##0"

        g_cell = ws.cell(r, 7)
        g_cell.value = f'=IFERROR(IF(D{r}="","",D{r}*F{r}),"")'
        g_cell.border = thin_border()
        g_cell.fill = fill(LIGHT_GOLD)
        g_cell.alignment = align("center")
        g_cell.font = font(bold=True, size=10)
        g_cell.number_format = "#,##0"

    # ── Row 20：小計 ──
    total_r = ITEM_START + ITEM_ROWS   # = 20
    ws.row_dimensions[total_r].height = 22
    ws.merge_cells(f"A{total_r}:F{total_r}")
    ws[f"A{total_r}"].value = "小　計"
    ws[f"A{total_r}"].font = font(bold=True, size=11, color="FFFFFF")
    ws[f"A{total_r}"].fill = fill("444444")
    ws[f"A{total_r}"].alignment = align("center", "center")
    ws[f"A{total_r}"].border = thin_border()
    ws[f"G{total_r}"].value = f"=SUM(G{ITEM_START}:G{ITEM_START+ITEM_ROWS-1})"
    ws[f"G{total_r}"].font = font(bold=True, size=11)
    ws[f"G{total_r}"].fill = fill(GRAY)
    ws[f"G{total_r}"].alignment = align("center", "center")
    ws[f"G{total_r}"].border = thin_border()
    ws[f"G{total_r}"].number_format = "#,##0"

    # ── Row 21：折扣 ──
    disc_r = total_r + 1   # = 21
    ws.row_dimensions[disc_r].height = 22
    ws.merge_cells(f"A{disc_r}:B{disc_r}")
    ws[f"A{disc_r}"].value = "折　扣"
    ws[f"A{disc_r}"].font = font(bold=True, size=11, color="FFFFFF")
    ws[f"A{disc_r}"].fill = fill("444444")
    ws[f"A{disc_r}"].alignment = align("center", "center")
    ws[f"A{disc_r}"].border = thin_border()

    ws[f"C{disc_r}"].value = None
    ws[f"C{disc_r}"].border = thin_border()
    ws[f"C{disc_r}"].fill = fill(LIGHT_GOLD)
    ws[f"C{disc_r}"].alignment = align("center")
    ws[f"C{disc_r}"].font = font(bold=True, size=11)
    ws[f"C{disc_r}"].number_format = '0"%"'

    ws[f"D{disc_r}"].value = "% OFF"
    ws[f"D{disc_r}"].border = thin_border()
    ws[f"D{disc_r}"].fill = fill(GRAY)
    ws[f"D{disc_r}"].alignment = align("center")
    ws[f"D{disc_r}"].font = font(size=9, color="666666")

    ws.merge_cells(f"E{disc_r}:F{disc_r}")
    ws[f"E{disc_r}"].value = "折扣金額"
    ws[f"E{disc_r}"].border = thin_border()
    ws[f"E{disc_r}"].fill = fill(GRAY)
    ws[f"E{disc_r}"].alignment = align("center")
    ws[f"E{disc_r}"].font = font(size=9, color="666666")

    ws[f"G{disc_r}"].value = f'=IF(C{disc_r}="","",-G{total_r}*C{disc_r}/100)'
    ws[f"G{disc_r}"].border = thin_border()
    ws[f"G{disc_r}"].fill = fill(GRAY)
    ws[f"G{disc_r}"].alignment = align("center")
    ws[f"G{disc_r}"].font = font(bold=True, color="CC0000", size=10)
    ws[f"G{disc_r}"].number_format = "#,##0"

    # ── Row 22：運費 ──
    ship_r = disc_r + 1   # = 22
    ws.row_dimensions[ship_r].height = 22
    ws.merge_cells(f"A{ship_r}:F{ship_r}")
    ws[f"A{ship_r}"].value = "運　費"
    ws[f"A{ship_r}"].font = font(bold=True, size=11, color="FFFFFF")
    ws[f"A{ship_r}"].fill = fill("444444")
    ws[f"A{ship_r}"].alignment = align("center", "center")
    ws[f"A{ship_r}"].border = thin_border()
    ws[f"G{ship_r}"].value = None
    ws[f"G{ship_r}"].border = thin_border()
    ws[f"G{ship_r}"].fill = fill(LIGHT_GOLD)
    ws[f"G{ship_r}"].alignment = align("center")
    ws[f"G{ship_r}"].font = font(bold=True, size=11)
    ws[f"G{ship_r}"].number_format = "#,##0"

    # ── Row 23：總計金額 ──
    final_r = ship_r + 1   # = 23
    ws.row_dimensions[final_r].height = 26
    ws.merge_cells(f"A{final_r}:F{final_r}")
    ws[f"A{final_r}"].value = "總計金額"
    ws[f"A{final_r}"].font = font(bold=True, size=12, color="FFFFFF")
    ws[f"A{final_r}"].fill = fill(DARK)
    ws[f"A{final_r}"].alignment = align("center", "center")
    ws[f"A{final_r}"].border = thin_border()
    ws[f"G{final_r}"].value = (
        f'=G{total_r}+IF(G{disc_r}="",0,G{disc_r})+IF(G{ship_r}="",0,G{ship_r})'
    )
    ws[f"G{final_r}"].font = font(bold=True, size=12, color=GOLD)
    ws[f"G{final_r}"].fill = fill(DARK)
    ws[f"G{final_r}"].alignment = align("center", "center")
    ws[f"G{final_r}"].border = thin_border()
    ws[f"G{final_r}"].number_format = "#,##0"

    # ── Row 24：備註 ──
    note_r = final_r + 1   # = 24
    ws.row_dimensions[note_r].height = 38
    _label(ws, f"A{note_r}", "備　　註")
    ws.merge_cells(f"B{note_r}:G{note_r}")
    ws[f"B{note_r}"].border = thin_border()
    ws[f"B{note_r}"].alignment = align("left", "top", wrap=True)
    ws[f"B{note_r}"].font = font(size=10)

    # ── Row 25：匯款資訊標題 ──
    bank_r = note_r + 1   # = 25
    ws.row_dimensions[bank_r].height = 18
    ws.merge_cells(f"A{bank_r}:G{bank_r}")
    ws[f"A{bank_r}"].value = "匯款資訊"
    ws[f"A{bank_r}"].font = font(bold=True, size=10, color="FFFFFF")
    ws[f"A{bank_r}"].fill = fill(DARK)
    ws[f"A{bank_r}"].alignment = align("center", "center")
    ws[f"A{bank_r}"].border = thin_border()

    # ── Rows 26~28：匯款明細 ──
    bank_fields = [
        ("銀　　行", "彰化銀行(009）南港科學園區分行"),
        ("戶　　名", "鉅鑫管理顧問有限公司"),
        ("帳　　號", "5383-01-056001-00"),
    ]
    for i, (label, value) in enumerate(bank_fields):
        r = bank_r + 1 + i
        ws.row_dimensions[r].height = 20
        lbl = ws.cell(r, 1, label)
        lbl.font = font(bold=True, size=9, color="FFFFFF")
        lbl.fill = fill("444444")
        lbl.alignment = align("center", "center")
        lbl.border = thin_border()
        ws.merge_cells(f"B{r}:G{r}")
        val = ws[f"B{r}"]
        val.value = value
        val.font = font(size=10, bold=(label == "帳　　號"))
        val.fill = fill(LIGHT_GOLD)
        val.alignment = align("left", "center")
        val.border = thin_border()

    # ── Row 29：警語 ──
    warn_r = bank_r + len(bank_fields) + 1   # = 29
    ws.row_dimensions[warn_r].height = 30
    ws.merge_cells(f"A{warn_r}:G{warn_r}")
    warn = ws[f"A{warn_r}"]
    warn.value = "⚠  應遠離火源、火花、火焰。切勿太陽光直射，以免高溫變質。"
    warn.font = Font(bold=False, size=9, color="8B4513", name="微軟正黑體")
    warn.fill = fill("FFF3CD")
    warn.alignment = align("center", "center", wrap=True)
    s = Side(style="medium", color=GOLD)
    warn.border = Border(left=s, right=s, top=s, bottom=s)

    # ── 列印設定：A4 Portrait，單頁適版 ──
    ws.print_area = f"A1:G{warn_r}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.5, header=0.2, footer=0.2
    )

    # ── 下拉選單：客戶姓名（B4）──
    cust_count = len(CUSTOMERS) + 50
    dv_cust = DataValidation(
        type="list",
        formula1=f"客戶名單!$A$2:$A${cust_count}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv_cust)
    dv_cust.add(ws["B4"])

    # ── 下拉選單：品項名稱（每行）──
    prod_count = len(PRODUCTS) + 30
    for i in range(ITEM_ROWS):
        r = ITEM_START + i
        dv_prod = DataValidation(
            type="list",
            formula1=f"商品目錄!$B$2:$B${prod_count}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
        )
        ws.add_data_validation(dv_prod)
        dv_prod.add(ws.cell(r, 2))

    ws.freeze_panes = "A8"
    ws.sheet_view.showGridLines = False

# ── 銷售紀錄 ────────────────────────────────────────────────
def build_sales(ws):
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12

    headers = ["訂單編號", "訂單日期", "客戶姓名", "購買品項", "數量", "單位", "單價", "金額"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = font(bold=True, color="FFFFFF")
        cell.fill = fill(DARK)
        cell.alignment = align("center")
        cell.border = thin_border()
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

# ── 輔助函式 ──────────────────────────────────────────────
def _label(ws, cell_ref, text):
    c = ws[cell_ref]
    c.value = text
    c.font = font(bold=True, color="FFFFFF", size=11)
    c.fill = fill(DARK)
    c.alignment = align("center", "center")
    c.border = thin_border()

def _input(ws, cell_ref, value=""):
    c = ws[cell_ref]
    c.value = value
    c.border = thin_border()
    c.font = font()
    c.alignment = align("left")

if __name__ == "__main__":
    build()
